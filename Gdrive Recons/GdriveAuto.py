from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
import pandas as pd
import calendar
import os

# Adding users with access to Recons folder
Register = {
    "shadrach" : "/Volumes/GoogleDrive/.shortcut-targets-by-id/11AuEbRHdz7JbK0smeqY3_YieDwACDbiT",
    "mark": "G:/.shortcut-targets-by-id/11AuEbRHdz7JbK0smeqY3_YieDwACDbiT"
}

name = input("Enter Name: ")
day = input("Recons day: ")
month = int(input("Recons month: "))
year = "20" + input("Recons year: ")
basePath = Register[name] + "/Recons"
month_abbr = calendar.month_abbr[month]
full_day = "_"+day+"_"+month_abbr+"_"+year[-2:]

print(full_day)
reconsPath  = str(basePath) + "/" + str(year) + "/" + str(month_abbr.upper()) + "/" + "Recons_" + month_abbr + " " + day
print(f"Recons day is {day}/{month}/{year}")
files = os.listdir(reconsPath)

def get_column(keyword):
    for i in range(first_row, last_row):
        for j in range(first_col, last_col):
            if sheet[str(get_column_letter(j)) + str(i)].value == keyword:
                return(get_column_letter(j))


#### MTN KOWRI COLLECTIONS (DEBIT) OVA ###






#### MTN KOWRI COLLECTIONS (DEBIT) INT ####
try:
    mtnKRCollectionsPath = reconsPath + "/KOWRI/MTN/Collection"
    print(mtnKRCollectionsPath)
    files= os.listdir(mtnKRCollectionsPath)

except:
    mtnKRCollectionsPath = reconsPath + "/KOWRI/MTN/KR Debit"
    print(mtnKRCollectionsPath)
    files= os.listdir(mtnKRCollectionsPath)

print(files)
wb = load_workbook(mtnKRCollectionsPath + '/MTN KR Debit Metabase'+full_day+'.xlsx')
ws1 = wb.create_sheet("Duplicates")
ws1.title = 'Duplicates'
sheet = wb[wb.sheetnames[0]]
file = pd.read_excel(mtnKRCollectionsPath + '/MTN KR Debit Metabase'+full_day+'.xlsx')
match_INT = 'MTN KR Debit Metabase_12_Dec_22.xlsx'

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

MTN_KR_Debit_INT_Volume = (last_row + 1) - (first_row + 1)
INT_VOLUME15 = MTN_KR_Debit_INT_Volume
print('MTN_KR_Debit_KR_Volume: ' + str(MTN_KR_Debit_INT_Volume))

MTN_KR_Debit_INT_Sum = 0.00

amountColumn = get_column('Amount')
int_id_col = get_column('IntegratorTransId')
for i in range(first_row+1, last_row+1):
    MTN_KR_Debit_INT_Sum = MTN_KR_Debit_INT_Sum + float(sheet[amountColumn + str(i)].value)

print('MTN_KR_Debit_KR_Sum: ' + str(MTN_KR_Debit_INT_Sum))
INT_VALUE15 = MTN_KR_Debit_INT_Sum
wb.close()
# # ---------------------------------------- DUPLICATES -------------------------------------------------
list = []
duplicate = []
count = 1
counter = 0
duplicates_value_sum = 0.00
for id in range(first_row+1, last_row+1):
    if sheet[int_id_col+str(id)].value not in list:
        list.append(sheet[int_id_col+str(id)].value)
    else:
        data = file[file['IntegratorTransId'] == (sheet[int_id_col+str(id)].value)]
        duplicate.append((sheet[int_id_col+str(id)].value))
        with pd.ExcelWriter(mtnKRCollectionsPath + '/MTN KR Debit Metabase'+full_day+'.xlsx', mode='a', engine="openpyxl", if_sheet_exists='overlay') as writer:
            data.to_excel(writer, sheet_name='Duplicates', startrow=count)
            count += 3
wb.close()

wb = load_workbook(mtnKRCollectionsPath + '/MTN KR Debit Metabase'+full_day+'.xlsx')
wb.active = wb['Duplicates']
dup_sheet = wb['Duplicates']


first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row


dup_col = column_index_from_string(amountColumn) + 1
for number in range(first_row+2, last_row+1, 3):
    duplicates_value_sum = duplicates_value_sum + dup_sheet[str(get_column_letter(dup_col))+str(number)].value
    count += 3
    counter += 1


print(f"Number of duplicates: {counter}")
dup_sheet[get_column_letter(dup_col) + str(last_row + 4)].value = str(round(duplicates_value_sum, 2))
dup_sheet[get_column_letter(dup_col) + str(last_row + 5)].value = str(counter)

print(f"Duplicates: {duplicate}")


wb.close()
wb.save(mtnKRCollectionsPath + '/MTN KR Debit Metabase Test'+full_day+'.xlsx')



############################ MTN KOWRI DISBURSEMENT (CREDIT) ###############################################33####

# mtnKRCreditPath = reconsPath + "/KOWRI/MTN/KR CREDIT"
# files= os.listdir(mtnKRDebitPath)

# for file in dir_list:
#     if file == 'MTN KR Credit' + yesterday + '.xlsx':
#         fileFound = True
#         matchOVAFound = True
#         wb = load_workbook('MTN KR Credit' + yesterday + '.xlsx')
#         sheet = wb[wb.sheetnames[0]]
#         match_OVA = 'MTN KR Credit' + yesterday + '.xlsx'
#         # DEFINE MAX AND MIN COLUMNS AND ROWS
#         first_col = wb.active.min_column
#         last_col = wb.active.max_column
#         first_row = wb.active.min_row
#         last_row = wb.active.max_row

#         MTN_KR_Credit_OVA_Volume = (last_row + 1) - (first_row + 1)
#         OVA_VOLUME14 = MTN_KR_Credit_OVA_Volume
#         print('MTN_KR_Credit_OVA_Volume: ' + str(MTN_KR_Credit_OVA_Volume))

#         MTN_KR_Credit_OVA_Sum = 0.00

#         amountColumn = get_column('Amount')
#         ova_id_col = get_column('IntegratorTransId')
#         for i in range(first_row + 1, last_row + 1):
#             MTN_KR_Credit_OVA_Sum = MTN_KR_Credit_OVA_Sum + sheet[amountColumn + str(i)].value

#         print('MTN_KR_Credit_OVA_Sum: ' + str(MTN_KR_Credit_OVA_Sum))
#         OVA_VALUE14 = MTN_KR_Credit_OVA_Sum
#         wb.close()
#         break
# if fileFound is False:
#     OVA_VOLUME14 = 0
#     OVA_VALUE14 = 0
# #
# #
# #
# fileFound= False
# #
# # ############################################### MTN KR CREDIT INT ############################################
# for file in dir_list:
#     if file =='MTN KR Credit Metabase'+yesterday+'.xlsx':
#         fileFound = True
#         matchINTFound = True
#         wb = load_workbook('MTN KR Credit Metabase'+yesterday+'.xlsx')
#         ws1 = wb.create_sheet("Duplicates")
#         ws1.title = 'Duplicates'
#         sheet = wb[wb.sheetnames[0]]
#         file = pd.read_excel('MTN KR Credit Metabase'+yesterday+'.xlsx')
#         match_INT = 'MTN KR Credit Metabase'+yesterday+'.xlsx'

#         # DEFINE MAX AND MIN COLUMNS AND ROWS
#         first_col = wb.active.min_column
#         last_col = wb.active.max_column
#         first_row = wb.active.min_row
#         last_row = wb.active.max_row

#         MTN_KR_Credit_INT_Volume = (last_row + 1) - (first_row + 1)
#         INT_VOLUME14 = MTN_KR_Credit_INT_Volume
#         print('MTN_KR_CREDIT_INT_Volume: ' + str(MTN_KR_Credit_INT_Volume))

#         MTN_KR_Credit_INT_Sum = 0.00

#         amountColumn = get_column('Amount')
#         int_id_col = get_column('External Transaction Id')
#         for i in range(first_row+1, last_row+1):
#             MTN_KR_Credit_INT_Sum = MTN_KR_Credit_INT_Sum + (-1 * sheet[amountColumn + str(i)].value)

#         print('MTN_KR_CREDIT_INT_Sum: ' + str(MTN_KR_Credit_INT_Sum))
#         INT_VALUE14 = MTN_KR_Credit_INT_Sum

#         # -------------------------------------------- Duplicates -----------------------------------------------
#         list = []
#         duplicate = []
#         count = 1
#         counter = 0
#         duplicates_value_sum = 0.00
#         for id in range(first_row+1, last_row+1):
#             if sheet[int_id_col+str(id)].value not in list:
#                 list.append(sheet[int_id_col+str(id)].value)
#             else:
#                 data = file[file['External Transaction Id'] ==(sheet[int_id_col+str(id)].value) ]
#                 duplicate.append((sheet[int_id_col+str(id)].value))
#                 with pd.ExcelWriter('MTN KR Credit Metabase'+yesterday+'.xlsx', mode='a', engine="openpyxl", if_sheet_exists='overlay') as writer:
#                     data.to_excel(writer, sheet_name='Duplicates', startrow=count)
#                     count += 3
#                     counter +=1
#         wb.close()
#         wb.save('MTN KR Credit Metabase'+yesterday+'.xlsx')

#         wb = load_workbook('MTN KR Credit Metabase'+yesterday+'.xlsx')
#         wb.active = wb['Duplicates']
#         dup_sheet = wb['Duplicates']

#         first_col = wb.active.min_column
#         last_col = wb.active.max_column
#         first_row = wb.active.min_row
#         last_row = wb.active.max_row

#         dup_col = column_index_from_string(amountColumn) + 1
#         for number in range(first_row+2, last_row+1, 3):
#             duplicates_value_sum = duplicates_value_sum + dup_sheet[get_column_letter(dup_col) + str(number)].value
#             count += 3
#             counter += 1
#             print(duplicates_value_sum)


#         print(f"Number of duplicates: {counter}")
#         dup_sheet[get_column_letter(dup_col) + str(last_row + 4)].value = str(round(duplicates_value_sum, 2))
#         dup_sheet[get_column_letter(dup_col) + str(last_row + 5)].value = str(counter)

#         print(f"Duplicates: {duplicate}")


#         wb.close()
#         wb.save('MTN KR Credit Metabase'+yesterday+'.xlsx')

#         if matchOVAFound is True and matchINTFound is True:
#             # ###################################### MISSING INT TRANSACTIONS ################################################
#             Owb = load_workbook(match_OVA)
#             Osheet = Owb[Owb.sheetnames[0]]

#             Iwb = load_workbook(match_INT)
#             Isheet = Iwb[Iwb.sheetnames[0]]
#             Iwb.active = Iwb[Iwb.sheetnames[0]]

#             Ofirst_col = Owb.active.min_column
#             Olast_col = Owb.active.max_column
#             Ofirst_row = Owb.active.min_row
#             Olast_row = Owb.active.max_row

#             Ifirst_col = Iwb.active.min_column
#             Ilast_col = Iwb.active.max_column
#             Ifirst_row = Iwb.active.min_row
#             Ilast_row = Iwb.active.max_row

#             missing_INT_list = []
#             missing_OVA_list = []
#             matchFound = False

#             for i in range(Ofirst_row+1, Olast_row+1):
#                 for j in range(Ifirst_row+1, Ilast_row+1):
#                     if Osheet[ova_id_col + str(i)].value == Isheet[int_id_col+str(j)].value:
#                         matchFound = True
#                         break
#                     else:
#                         matchFound = False
#                 if matchFound == False:
#                     missing_INT_list.append(Osheet[ova_id_col + str(i)].value)
#             print(f"Missing integrator transactions: {missing_INT_list}")
#             ws2 = Iwb.create_sheet('Missing Integrator Transactions')
#             ws2.title = 'Missing Integrator Transactions'
#             file = pd.read_excel(match_OVA)
#             count = 0
#             for transaction in missing_INT_list:
#                 data = file[file['IntegratorTransId'] == transaction]
#                 with pd.ExcelWriter(match_INT, mode='a', engine="openpyxl", if_sheet_exists='overlay') as writer:
#                     data.to_excel(writer, sheet_name='Missing Integrator Transactions', startrow=count)
#                     count += 2
#             Iwb.close()
#             Owb.close()

#             # ######################################### MISSING OVA TRANSACTIONS ##############################################
#             Owb = load_workbook(match_OVA)
#             Osheet = Owb[Owb.sheetnames[0]]

#             Iwb = load_workbook(match_INT)
#             file = pd.read_excel(match_INT)
#             Isheet = Iwb[Iwb.sheetnames[0]]
#             Iwb.active = Iwb[Iwb.sheetnames[0]]

#             Ofirst_col = Owb.active.min_column
#             Olast_col = Owb.active.max_column
#             Ofirst_row = Owb.active.min_row
#             Olast_row = Owb.active.max_row

#             Ifirst_col = Iwb.active.min_column
#             Ilast_col = Iwb.active.max_column
#             Ifirst_row = Iwb.active.min_row
#             Ilast_row = Iwb.active.max_row

#             missing_INT_list = []
#             missing_OVA_list = []
#             matchFound = False

#             for i in range(Ifirst_row + 1, Ilast_row + 1):
#                 for j in range(Ofirst_row + 1, Olast_row + 1):
#                     if Isheet[int_id_col + str(i)].value == Osheet[ova_id_col + str(j)].value:
#                         matchFound = True
#                         break
#                     else:
#                         matchFound = False
#                 if matchFound == False:
#                     missing_OVA_list.append(Isheet[int_id_col + str(i)].value)
#             print(f"Missing OVA transactions: {missing_OVA_list}")
#             ws3 = wb.create_sheet('Missing OVA Transactions')
#             ws3.title = 'Missing OVA Transactions'
#             count = 0
#             for transaction in missing_OVA_list:
#                 data = file[file['External Transaction Id'] == transaction]
#                 with pd.ExcelWriter(match_INT, mode='a', engine="openpyxl", if_sheet_exists='overlay') as writer:
#                     data.to_excel(writer, sheet_name='Missing OVA Transactions', startrow=count)
#                     count += 2
#             Iwb.close()
#             Owb.close()
#         break

# if fileFound is False:
#     INT_VOLUME14 = 0
#     INT_VALUE14 = 0
# #
# #
# fileFound = False
# matchOVAFound = False
# matchINTFound = False

#### VODAFONE KOWRI COLLECTIONS (CashIn) ####



#### VODAFONE KOWRI DISBURSEMNT (CashOut) ####
