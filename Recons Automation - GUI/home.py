import customtkinter as ctk
import tkinter as tk
from tkinter import ttk
import calendar
from datetime import datetime,date,timedelta
from PIL import Image, ImageTk
import sys, os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
import pandas as pd
import warnings

warnings.filterwarnings("ignore")


class Console(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.text_widget = tk.Text(self, wrap=tk.WORD)
        self.text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.scrollbar = tk.Scrollbar(self, command=self.text_widget.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.text_widget.config(yscrollcommand=self.scrollbar.set)
        
        sys.stdout = self
        
    def write(self, text):
        self.text_widget.insert(tk.END, text)
        self.text_widget.see(tk.END)  # Automatically scroll to the end
        
    def flush(self):
        pass


def main ():
    root = ctk.CTk()
    root.title("KowRecons")
    root.option_add("*tearOff", False)
    root._set_appearance_mode("light")
    root.geometry("500x550")


    def render_recons_frame(tab):
        recons_frame = ctk.CTkFrame(tab,fg_color="white")

        logo_img = ctk.CTkImage(Image.open("C:\\Users\\Mark\\repos\DreamOval_Projects\Recons Automation - GUI\\assets\\KowriLogo.png"),size=(200,70))
        label = ctk.CTkLabel(recons_frame, image=logo_img,text="",fg_color="white")
        label.pack()
        


        yesterday_frame = ttk.LabelFrame(recons_frame,text="Set Recons Date",width=500,height=400)
        yesterday_frame.pack()
        yesterday_frame.pack_propagate(False)
        
        yesterady_msg = ctk.CTkLabel(yesterday_frame, text="Recons for yesterday?",text_color="green")
        yesterady_msg.pack()

        recons_var = ctk.StringVar()
        recons_var.set("yes")



        def yesterdayRecons():
            dayCombo.configure(state="disabled")
            monthCombo.configure(state="disabled")
            yearCombo.configure(state="disabled")
            today = datetime.today().strftime('%d_%b_%y')

            # Yesterday's date
            yesterday = (datetime.now() - timedelta(1)).strftime('_%d %b_%y')
            print(yesterday)

            # Current month
            current_month = calendar.month_abbr[date.today().month].upper()
            print(current_month)

            recons_yesterday = str((datetime.now() - timedelta(1)).strftime('%Y-%m-%d') + ' 00:00:00')
            print(recons_yesterday)

            if date.today().month < 10:
                GIPmonth = "0" + str(date.today().month)
            else:
                GIPmonth = str(date.today().month)

            if date.today().day < 10:
                GIPday = "0" + str(date.today().day + 1)
            else:
                GIPday = str(date.today().day + 1)

            GIPdate = str(str(date.today().year) + str(GIPmonth) + str((GIPday)))
            print(GIPdate)

        def customDateRecons():
                dayCombo.configure(state="enabled")
                monthCombo.configure(state="enabled")
                yearCombo.configure(state="enabled")


        def displayValues():
            if recons_var.get() == "no":
                customDay = dayCombo.get()
                customMonth = monthCombo.get()
                customYear = yearCombo.get()[-2:]
                GIPday = int(customDay) + 1
                if int(customDay) < 10:
                    customDay = "0" + customDay
                month_name = calendar.month_abbr[int(customMonth)]
                if int(customMonth) < 10:
                    customMonth = "0" + customMonth
                yesterday = "_" + customDay + " " + month_name + "_" + customYear
                print(yesterday)

                recons_yesterday = "20" + customYear + "-" + customMonth + "-" + customDay + ' 00:00:00'
                print(recons_yesterday)

                if int(customMonth) < 10:
                    GIPmonth = customMonth

                if int(GIPday) < 10:
                    GIPday = "0" + str(GIPday)
                GIPdate = str("20" + str(customYear) + str(GIPmonth) + str((GIPday)))
                print(GIPdate)

            elif recons_var.get() == "yes":
                yesterdayRecons()

            path = "C:\\Users\Mark\\repos\DreamOval_Projects\\Recons Automation - GUI"
            dir_list = os.listdir(path=path)
            print(dir_list)
            # Check if metabase file is available
            metaFileFound = False
            for file in dir_list:
                if file == 'Metabase' + yesterday + '.xlsx':
                    metaFileFound = True
                    break
                else:
                    metaFileFound = False
            def get_column(keyword):
                for i in range(first_row, last_row + 1):
                    for j in range(first_col, last_col + 1):
                        if sheet[str(get_column_letter(j)) + str(i)].value == keyword:
                            return (get_column_letter(j))
            #  ######################################## SLYDEPULL PROMPTS OVA #####################################################
            for file in dir_list:
                if file == 'MTN Prompt' + yesterday + '.xlsx':
                    fileFound = True
                    matchOVAFound = True
                    Prompt_wb = load_workbook(f'{path}\\MTN Prompt' + yesterday + '.xlsx')
                    sheet = Prompt_wb[Prompt_wb.sheetnames[0]]

                    # DEFINE MAX AND MIN COLUMNS AND ROWS
                    first_col = Prompt_wb.active.min_column
                    last_col = Prompt_wb.active.max_column
                    first_row = Prompt_wb.active.min_row
                    last_row = Prompt_wb.active.max_row

                    MTN_PROMPT_OVA_Volume = (last_row + 1) - (first_row + 1)
                    OVA_VOLUME2 = MTN_PROMPT_OVA_Volume
                    print('MTN_PROMPT_OVA_Volume: ' + str(MTN_PROMPT_OVA_Volume))

                    MTN_PROMPT_OVA_Sum = 0.00

                    amountColumn = get_column('Amount')
                    ova_id_col = get_column('External Transaction Id')
                    tmp_id_col = ova_id_col
                    alternate_id_col = get_column("Id")
                    for i in range(first_row + 1, last_row + 1):
                        check = sheet[amountColumn + str(i)].value
                        MTN_PROMPT_OVA_Sum = MTN_PROMPT_OVA_Sum + abs(float(sheet[amountColumn + str(i)].value))

                    print('MTN_PROMPT_OVA_Sum: ' + str(MTN_PROMPT_OVA_Sum))
                    OVA_VALUE2 = MTN_PROMPT_OVA_Sum
                    Prompt_wb.close()
                    Prompt_wb.save(f'{path}\\MTN Prompt' + yesterday + '- Recons.xlsx')
                    match_OVA = f'{path}\\MTN Prompt' + yesterday + '- Recons.xlsx'

                    wb = load_workbook(match_OVA)
                    ws1 = wb.create_sheet("Duplicates")
                    ws1.title = 'Duplicates'

                    ws1 = wb.create_sheet("Missing Integrator Transactions")
                    ws1.title = 'Missing Integrator Transactions'

                    ws1 = wb.create_sheet("Missing Integrator Transactions")
                    ws1.title = 'Missing Integrator Transactions'

                    wb.close()
                    wb.save(match_OVA)
                    break
                else:
                    fileFound = False

            if fileFound is False:
                OVA_VOLUME2 = 0
                OVA_VALUE2 = 0
            #
            #
            #

            # ################################### MTN(SLYDEPULL) PROMPT INT ###############################################
            if metaFileFound is True:
                matchINTFound = True
                wb = load_workbook(f'{path}\\Metabase' + yesterday + '.xlsx')
                file = pd.read_excel(f'{path}\\Metabase' + yesterday + '.xlsx', sheet_name='Query result')

                sheet = wb['Query result']
                wb.active = wb['Query result']
                first_col = wb.active.min_column
                last_col = wb.active.max_column
                first_row = wb.active.min_row
                last_row = wb.active.max_row

                int_id_col = get_column('integratorTransId')
                transIdHeader = "integratorTransId"
                if int_id_col is None:
                    int_id_col = get_column("IntegratorTransId")
                    transIdHeader = "IntegratorTransId"
                debitCreditFlagColumn = get_column('creditDebitFlag')
                creditDebitHeader = "creditDebitFlag"
                if debitCreditFlagColumn is None:
                    debitCreditFlagColumn = get_column("CreditDebitFlag")
                    creditDebitHeader = "CreditDebitFlag"
                serviceName_col = get_column('serviceName')
                serviceNameHeader = "serviceName"
                if serviceName_col is None:
                    serviceName_col = get_column("ServiceName")
                    serviceNameHeader = "ServiceName"
                amountColumn = get_column("amount")
                if amountColumn is None:
                    amountColumn = get_column("Amount")

                INT_VOLUME2 = 0.00
                INT_VALUE2 = 0.00
                row_count = 1
                count = 0
                sum = 0.00
                counter = 0
                headerChecker = True
                # ------------------------------------------------- Duplicates ---------------------------------------------------------
                list = []
                duplicates = []
                duplicates_value_sum= 0.00
                for i in range(first_row + 1, last_row + 1):
                    if sheet[serviceName_col + str(i)].value == 'MTN OVA' and sheet[debitCreditFlagColumn + str(i)].value == 'C':
                        count += 1
                        sum = sum + abs(float(sheet[amountColumn + str(i)].value))
                        if sheet[int_id_col + str(i)].value not in list:
                            list.append(sheet[int_id_col + str(i)].value)
                        elif sheet[int_id_col + str(i)].value not in duplicates:
                                data = file[(file[transIdHeader] == sheet[int_id_col + str(i)].value) & (file[serviceNameHeader] == 'MTN OVA') & (file[creditDebitHeader] == 'C')]
                                try:
                                    data_set_sum = data.iloc[1:, data.columns.get_loc("Amount")].sum()
                                except:
                                    data_set_sum = data.iloc[1:, data.columns.get_loc("amount")].sum()
                                duplicates_value_sum = duplicates_value_sum + data_set_sum
                                duplicates.append(sheet[int_id_col + str(i)].value)
                                if counter > 0:
                                    headerChecker = False
                                with pd.ExcelWriter(match_OVA, mode='a', engine="openpyxl", if_sheet_exists='overlay') as writer:
                                    data.to_excel(writer, sheet_name='Duplicates', startrow=row_count, header=headerChecker)
                                    if headerChecker is False:
                                        row_count += 2
                                    else:
                                        row_count += (data.shape[0]+1)
                                    counter += 1
                        else:
                            if sheet[int_id_col + str(i)].value in duplicates:
                                counter += 1
                            continue
                    else:
                        continue
                INT_VOLUME2 = count
                INT_VALUE2 = sum
                print(f"MTN Prompt INT Volume: {count}")
                print(f"MTN Prompt INT Value: {sum}")

                wb.close()
                Mwb = load_workbook(match_OVA)

                wb = load_workbook(match_OVA)
                wb.active = wb['Duplicates']
                dup_sheet = wb['Duplicates']
                sheet = dup_sheet

                first_col = wb.active.min_column
                last_col = wb.active.max_column
                first_row = wb.active.min_row
                last_row = wb.active.max_row

                if counter > 0:
                    amountColumn = get_column('Amount')
                    if amountColumn is None:
                        amountColumn = get_column("amount")

                    print(f"Number of duplicates: {counter}")
                    dup_sheet[amountColumn + str(last_row + 4)].value = str(round(duplicates_value_sum, 2))
                    dup_sheet[amountColumn + str(last_row + 5)].value = str(counter)

                    print(f"Duplicates: {duplicates}")

                    DUPLICATES_VOLUME2 = counter
                    DUPLICATES_VALUE2 = duplicates_value_sum
                else:
                    DUPLICATES_VOLUME2 = 0
                    DUPLICATES_VALUE2 = 0

                wb.close()
                wb.save(match_OVA)

                if matchOVAFound is True and matchINTFound is True:
                    # --------------------------------------- MISSING INT TRANSACTIONS ------------------------------------------------
                    Owb = load_workbook(match_OVA)
                    Osheet = Owb[Owb.sheetnames[0]]
                    Owb.active = Owb[Owb.sheetnames[0]]
                    file = pd.read_excel(match_OVA)
                    sheet = Osheet

                    Iwb = load_workbook(f'{path}\\Metabase' + yesterday + '.xlsx')
                    Isheet = Iwb['Query result']
                    Iwb.active = Iwb['Query result']

                    Ofirst_col = Owb.active.min_column
                    Olast_col = Owb.active.max_column
                    Ofirst_row = Owb.active.min_row
                    Olast_row = Owb.active.max_row

                    Ifirst_col = Iwb.active.min_column
                    Ilast_col = Iwb.active.max_column
                    Ifirst_row = Iwb.active.min_row
                    Ilast_row = Iwb.active.max_row

                    missing_INT_list = []
                    missing_OVA_list = []
                    matchFound = False
                    headerChecker = True
                    counter = 0

                    for i in range(Ofirst_row + 1, Olast_row + 1):
                        for j in range(Ifirst_row + 1, Ilast_row + 1):
                            if Osheet[ova_id_col + str(i)].value is None or Osheet[ova_id_col + str(i)].value == "#VALUE!":
                                ova_id_col = alternate_id_col
                                matchFound = False
                                break
                            if (Isheet[serviceName_col + str(j)].value == 'MTN OVA') and (
                                    Isheet[debitCreditFlagColumn + str(j)].value == 'C'):
                                if str(Osheet[ova_id_col + str(i)].value).lstrip('=') in str(Isheet[int_id_col + str(j)].value):
                                    matchFound = True
                                    break
                                else:
                                    matchFound = False
                        if matchFound is False:
                            missing_INT_list.append(Osheet[ova_id_col + str(i)].value)
                            ova_id_col = tmp_id_col

                    print(f"Missing integrator transactions: {missing_INT_list}")

                    count = 0
                    for transaction in missing_INT_list:
                        data = (file[file['External Transaction Id'] == transaction])
                        if data.empty:
                            data = (file[file['Id'] == transaction])
                        if counter > 0:
                            headerChecker = False
                        with pd.ExcelWriter(match_OVA, mode='a', engine="openpyxl", if_sheet_exists='overlay') as writer:
                            data.to_excel(writer, sheet_name='Missing Integrator Transactions', startrow=count,
                                        header=headerChecker)
                            if headerChecker is False:
                                count += 1
                            else:
                                count += 2
                            counter += 1

                    Iwb.close()
                    Owb.close()
                    Mwb.close()

                    # -------------------------------------- MISSING OVA TRANSACTIONS --------------------------------------------------
                    Owb = load_workbook(match_OVA)
                    Osheet = Owb[Owb.sheetnames[0]]
                    Owb.active = Owb[Owb.sheetnames[0]]
                    sheet = Osheet
                    Mwb = load_workbook(match_OVA)

                    Iwb = load_workbook(f'{path}\\Metabase' + yesterday + '.xlsx')
                    file = pd.read_excel(f'{path}\\Metabase' + yesterday + '.xlsx', sheet_name='Query result')
                    Isheet = Iwb['Query result']
                    Iwb.active = Iwb['Query result']

                    Ofirst_col = Owb.active.min_column
                    Olast_col = Owb.active.max_column
                    Ofirst_row = Owb.active.min_row
                    Olast_row = Owb.active.max_row

                    Ifirst_col = Iwb.active.min_column
                    Ilast_col = Iwb.active.max_column
                    Ifirst_row = Iwb.active.min_row
                    Ilast_row = Iwb.active.max_row

                    missing_INT_list = []
                    missing_OVA_list = []
                    matchFound = False
                    headerChecker = True
                    counter = 0

                    for i in range(Ifirst_row + 1, Ilast_row + 1):
                        if (Isheet[serviceName_col + str(i)].value == 'MTN OVA') and (
                                Isheet[debitCreditFlagColumn + str(i)].value == 'C'):
                            for j in range(Ofirst_row + 1, Olast_row + 1):
                                if str(Osheet[ova_id_col + str(j)].value).lstrip('=') in str(Isheet[int_id_col + str(i)].value):
                                    matchFound = True
                                    break
                                else:
                                    matchFound = False
                            if matchFound is False:
                                missing_OVA_list.append(Isheet[int_id_col + str(i)].value)
                        else:
                            continue

                    print(f"Missing OVA transactions: {missing_OVA_list}")

                    count = 0
                    for transaction in missing_OVA_list:
                        try:
                            data = file[(file['integratorTransId'] == str(transaction))]
                        except KeyError:
                            data = file[(file['IntegratorTransId'] == str(transaction))]
                        if counter > 0:
                            headerChecker = False
                        with pd.ExcelWriter(match_OVA, mode='a', engine="openpyxl", if_sheet_exists='overlay') as writer:
                            data.to_excel(writer, sheet_name='Missing OVA Transactions', startrow=count, header=headerChecker)
                            if headerChecker is False:
                                count += 1
                            else:
                                count += 2
                            counter += 1

                    Iwb.close()
                    Owb.close()
                    Mwb.close()
            else:
                INT_VOLUME2 = 0
                INT_VALUE2 = 0
                DUPLICATES_VOLUME2 = 0
                DUPLICATES_VALUE2 = 0

            #
            #
            OVA_VOLUME3 = 0
            OVA_VALUE3 = 0
            INT_VOLUME3 = 0
            INT_VALUE3 = 0
            DUPLICATES_VALUE3 = 0
            DUPLICATES_VOLUME3 = 0
            #
            #
            #
            matchOVAFound = False
            matchINTFound = False
            fileFound = False

            
        def topLevelConsole():
            consoleWindow = tk.Toplevel(root)
            consoleWindow.title("Running")
            consoleWindow.geometry("500x650")
            consoleWindow.attributes('-topmost', True)

            consoleWindow.columnconfigure(index=0, weight=1)
            consoleWindow.columnconfigure(index=1, weight=1)
            consoleWindow.columnconfigure(index=2, weight=1)
            consoleWindow.rowconfigure(index=0, weight=1)
            consoleWindow.rowconfigure(index=1, weight=1)
            consoleWindow.rowconfigure(index=2, weight=1)

            console_frame = ttk.Frame(consoleWindow)
            console_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)


            console = Console(console_frame)
            console.pack(fill=tk.BOTH, expand=True)



            
            
    
        day_value_list = []
        for i in range(1,32):
            day_value_list.append(i)

        month_value_list = []
        for i in range(1,13):
            month_value_list.append(i)

        
        yesRadio = ttk.Radiobutton(yesterday_frame, text="Yes", variable=recons_var, value="yes",command=yesterdayRecons)
        yesRadio.pack(anchor=ctk.W)
        noRadio = ttk.Radiobutton(yesterday_frame, text="No", variable=recons_var, value="no",command=customDateRecons)
        noRadio.pack(anchor=ctk.W)


        dayLabel = ctk.CTkLabel(yesterday_frame,text="Day")
        dayLabel.pack()
        dayCombo = ttk.Combobox(yesterday_frame, state="disabled", values=day_value_list)
        dayCombo.current(0)
        dayCombo.pack()

        monthLabel = ctk.CTkLabel(yesterday_frame,text="Month")
        monthLabel.pack()
        monthCombo = ttk.Combobox(yesterday_frame, state="disabled", values=month_value_list)
        monthCombo.current(0)
        monthCombo.pack()

        yearLabel = ctk.CTkLabel(yesterday_frame,text="Year")
        yearLabel.pack()
        yearCombo = ttk.Combobox(yesterday_frame, state="disabled", values=["2022","2023"])
        yearCombo.current(0)
        yearCombo.pack()

        startButton = ctk.CTkButton(yesterday_frame,text="Start",fg_color="green",hover_color="#1bcf48",command=lambda:(startProgress(),topLevelConsole(),displayValues()))
        startButton.pack(pady=(20,0))


        progress_bar = ttk.Progressbar(yesterday_frame, mode="indeterminate")
        progress_bar.pack_forget()
        # progress_bar.pack()

        def startProgress():
            root.update()
            progress_bar.pack()
            progress_bar.start()

        
        recons_frame.pack(fill="both", expand=True)

    
    # Make the window interactive
    root.columnconfigure(index=0, weight=1)
    root.columnconfigure(index=1, weight=1)
    root.columnconfigure(index=2, weight=1)
    root.rowconfigure(index=0, weight=1)
    root.rowconfigure(index=1, weight=1)
    root.rowconfigure(index=2, weight=1)

    style = ttk.Style(root)
    # Import the tcl file
    root.tk.call("source", "Forest-ttk-theme-master/forest-light.tcl")

    # Set the theme with the theme_use method
    style.theme_use("forest-light")


    tab_control = ttk.Notebook(root)

    home_tab = ttk.Frame(tab_control)
    recons_tab = ttk.Frame(tab_control)
    logs_tab = ttk.Frame(tab_control)

    tab_control.add(home_tab, text="Home")
    tab_control.add(recons_tab, text="Recons")
    tab_control.add(logs_tab, text="Logs")

    render_recons_frame(recons_tab)

    tab_control.pack(side=ctk.LEFT, fill="both", expand=True)

    root.mainloop()

if __name__ == "__main__":
    main()