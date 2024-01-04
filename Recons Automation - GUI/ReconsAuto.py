import customtkinter as ctk
import tkinter as tk
from tkinter import ttk
from typing import Literal, Tuple
import calendar
from datetime import datetime, date, timedelta
from PIL import Image, ImageTk
import sys, os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
import pandas as pd
import warnings, time, threading, subprocess

warnings.filterwarnings("ignore")


#################################### ATTEMPTS FOR DYNAMIC DIRECTORY (INVESTIGATE LATER)#####################################
# if getattr(sys, 'frozen', False):
#     EXE_LOCATION = os.path.dirname( sys.executable ) # cx_Freeze frozen
# else:
#     EXE_LOCATION = os.path.dirname( os.path.realpath( __file__ ) ) # Other packers

# my_logo = os.path.join( EXE_LOCATION, "assets", "KowriLogo.png" )
# my_theme = os.path.join( EXE_LOCATION, "Forest-ttk-theme-master", "forest-light.tcl" )

# curpath = os.getcwd()
# logoPath = os.path.abspath("assets/KowriLogo.png")
############################################################################################################################


class Console(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.text_widget = tk.Text(self, wrap=tk.WORD)
        self.text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.scrollbar = tk.Scrollbar(self, command=self.text_widget.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.text_widget.config(yscrollcommand=self.scrollbar.set)

        sys.stdout = self
        sys.stderr = self

    def write(self, text):
        self.text_widget.insert(tk.END, text)
        self.text_widget.see(tk.END)  # Automatically scroll to the end

    def flush(self):
        pass


def restart_program():
    python = sys.executable  # Get the s to the Python interpreter
    subprocess.Popen(
        [python] + sys.argv
    )  # Restart the program with the same command line arguments
    os._exit(0)  # Forcefully exit the current instance


home_directory = os.path.expanduser("~")
user = os.path.basename(home_directory)


warnings.filterwarnings("ignore")
#################################### DYNAMIC DIRECTORY FIX (WORKING)#############################################

# def get_script_directory():
#     # Determine the directory where the script is located
#     if getattr(sys, 'frozen', False):
#         # We are running from a bundled executable
#         return os.path.dirname(sys.executable)
#     else:
#         # We are running the script directly
#         return os.path.dirname(os.path.abspath(__file__))

# # Get the script directory
# script_dir = get_script_directory()
# os.chdir(script_dir)
# print(script_dir)

# # Use script_dir as the reference point for accessing assets and files
# asset_path = os.path.join(script_dir, 'assets', 'example.txt')
###################################################################################################################


def set_working_directory_to_script_location():
    if getattr(sys, "frozen", False):
        # We are running from a bundled executable
        script_dir = os.path.dirname(sys.executable)
    else:
        # We are running the script directly
        script_dir = os.path.dirname(os.path.abspath(__file__))

    os.chdir(script_dir)
    return script_dir


# Call the function to set the working directory
script_dir = set_working_directory_to_script_location()


def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS2
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


tx_id_col_names = [
    "integratorTransId",
    "IntegratorTransId",
    "Transaction Id",
    "TransId",
    "External Transaction Id",
    "BillerTransId",
    "Id",
    "External Payment Request â†’ Institution Trans ID",
    "Merchant Transaction Reference",
    "REMARKS2",
    "External Payment Request → Institution Trans ID",
    "Order Code",
    "Order ID",
    "Integrator Trans ID",
]
amount_col_names = [
    "Amount",
    "amount",
    "Paid in",
    "Paid In",
    "Withdrawn",
    "AMOUNT",
    "Amount ($)",
    "Actual Amount ($)",
    "Transaction Amount (GHC.)",
    "TRANSACTION_AMOUNT",
    "Transaction Amount (amount only)",
    "Order Amount (amount only)",
    "Real Total",
]


def check_for_file(file_name):
    # This function is called when calling the run_recons fucntion.
    # It checks if a particular file can be found in the directory and returns the file name else returns None
    if file_name in os.listdir():
        return file_name
    else:
        return None


def get_write_double_ova_val(
    ova_file: str or None,
    num_lines_of_header: Tuple[int, int],
    alt_recons_name: str,
    file_output_name: str,
):
    ova_df1 = pd.read_excel(f"{script_dir}/{ova_file}", skiprows=num_lines_of_header[0])

    ova_volume = len(ova_df1)
    ova_volumes[21] = ova_volume
    recons_file = (
        f"{alt_recons_name} - Recons.xlsx"
        if ova_file is None
        else f"{ova_file[:-5]} - Recons.xlsx"
    )
    df1_amount_col = ""
    for name in amount_col_names:
        if name in ova_df1.columns:
            if not ova_df1[name].isna().all():
                df1_amount_col = name
                break

    ova_value = ova_df1[df1_amount_col].abs().sum()
    ova_values[21] = ova_value
    print(f"{file_output_name} OVA VOLUME : {ova_volume}")
    print(f"{file_output_name} OVA VALUE : {ova_value}")

    with pd.ExcelWriter(
        recons_file, engine="openpyxl", mode="w"
    ) as writer:  # specify new file name to write to
        ova_df1.to_excel(
            writer, sheet_name="Sheet1", index=False
        )  # save original data into first sheet of new file
        return recons_file


def get_write_double_int_val(
    int_files: Tuple[str or None, str or None],
    num_lines_of_header: Tuple[int, int],
    alt_recons_name: str,
    file_output_name: str,
    recons_file: str,
):
    # see get_write_double_ova_val function
    try:
        int_df1 = pd.read_excel(
            f"{script_dir}/{int_files[0]}", skiprows=num_lines_of_header[0]
        )
        int_df2 = pd.read_excel(
            f"{script_dir}/{int_files[1]}", skiprows=num_lines_of_header[1]
        )
        int_df2 = int_df2.loc[int_df2["Status"] == "CONFIRMED"]

        int_volume = len(int_df1) + len(int_df2)

        df1_amount_col = ""
        for name in amount_col_names:
            if name in int_df1.columns:
                if not int_df1[name].isna().all():
                    df1_amount_col = name
                    break

        df2_amount_col = ""
        for name in amount_col_names:
            if name in int_df2.columns:
                if not int_df2[name].isna().all():
                    df2_amount_col = name
                    break

        int_value = (
            int_df1[df1_amount_col].abs().sum() + int_df2[df2_amount_col].abs().sum()
        )
        print(f"{file_output_name} INT VOLUME : {int_volume}")
        print(f"{file_output_name} INT VALUE : {int_value}")
        dup, dup_val = find_duplicates(int_df1)
        write_duplicate_data(
            amount_col_name=df1_amount_col, df=dup, value=dup_val, file_name=recons_file
        )
        dup2, dup_val = find_duplicates(int_df2)
        write_duplicate_data(
            amount_col_name=df2_amount_col,
            df=dup2,
            value=dup_val,
            file_name=recons_file,
            last_row_index=len(dup) + 3,
        )
        return int_volume, int_value
    except ValueError:
        try:
            int_files = tuple(item for item in int_files if item is not None)
            if int_files[0] == f"MPGS KC{yesterday}.xlsx":
                int_df1 = pd.read_excel(
                    f"{script_dir}/{int_files[0]}", skiprows=num_lines_of_header[0]
                )
                int_volume = len(int_df1)

                amount_col = ""
                for name in amount_col_names:
                    if name in int_df1.columns:
                        if not int_df1[name].isna().all():
                            amount_col = name
                            break

                int_value = int_df1[amount_col].abs().sum()
                print(f"{file_output_name} INT VOLUME : {int_volume}")
                print(f"{file_output_name} INT VALUE : {int_value}")
                dup, dup_val = find_duplicates(int_df1)
                write_duplicate_data(
                    amount_col_name=amount_col,
                    df=dup,
                    value=dup_val,
                    file_name=recons_file,
                )
                return int_volume, int_value
            elif int_files[0] == f"MPGS_trn{yesterday}.xlsx":
                int_df1 = pd.read_excel(
                    f"{script_dir}/{int_files[0]}", skiprows=num_lines_of_header[0]
                )
                int_df1 = int_df1.loc[int_df1["Status"] == "CONFIRMED"]
                int_volume = len(int_df1)

                amount_col = ""
                for name in amount_col_names:
                    if name in int_df1.columns:
                        if not int_df1[name].isna().all():
                            amount_col = name
                            break

                int_value = int_df1[amount_col].abs().sum()
                print(f"{file_output_name} INT VOLUME : {int_volume}")
                print(f"{file_output_name} INT VALUE : {int_value}")
                dup, dup_val = find_duplicates(int_df1)
                write_duplicate_data(
                    amount_col_name=amount_col,
                    df=dup,
                    value=dup_val,
                    file_name=recons_file,
                )
                return int_volume, int_value
        except:
            print("INT FILE NOT FOUND")


def update_recons_sheet():
    fwb = load_workbook("Reconciliations 2023.xlsx")

    if datetime.today().day == 1:
        today = date.today()
        prev_month = (
            today.replace(month=today.month - 1)
            if today.month > 1
            else today.replace(month=12, year=today.year - 1)
        )
        prev_month_abbr = prev_month.strftime("%b").upper()
        fsheet = fwb[prev_month_abbr]
        fwb.active = fwb[prev_month_abbr]
    else:
        fsheet = fwb[current_month]
        fwb.active = fwb[current_month]

    first_col = fwb.active.min_column  # type: ignore
    last_col = fwb.active.max_column  # type: ignore
    first_row = fwb.active.min_row  # type: ignore
    last_row = fwb.active.max_row  # type: ignore

    start_row = 0
    for row in range(first_row + 1, last_row + 1):
        if str(fsheet["A" + str(row)].value) == recons_yesterday:
            start_row = row
            break

    for row in range(0, 11):
        fsheet["E" + str(start_row)].value = ova_volumes[row]
        fsheet["F" + str(start_row)].value = abs(ova_values[row])
        fsheet["G" + str(start_row)].value = int_volumes[row]
        fsheet["H" + str(start_row)].value = abs(int_values[row])
        fsheet["M" + str(start_row)].value = dup_volumes[row]
        fsheet["N" + str(start_row)].value = abs(dup_values[row])

        start_row += 1
    start_row += 10

    for row in range(21, 30):
        fsheet["E" + str(start_row)].value = ova_volumes[row]
        fsheet["F" + str(start_row)].value = abs(ova_values[row])
        fsheet["G" + str(start_row)].value = int_volumes[row]
        fsheet["H" + str(start_row)].value = abs(int_values[row])
        fsheet["M" + str(start_row)].value = dup_volumes[row]
        fsheet["N" + str(start_row)].value = abs(dup_values[row])
        start_row += 1
    fwb.close()
    fwb.save("Reconciliations 2023.xlsx")


def find_duplicates(int_df: pd.DataFrame):
    trans_id_col = ""
    for name in tx_id_col_names:
        if name in int_df.columns:
            trans_id_col = name
            break

    amount_col = ""
    for name in amount_col_names:
        if name in int_df.columns:
            if not int_df[name].isna().all():
                amount_col = name
                break

    if trans_id_col == "":
        raise Exception("No transaction id column found")

    unique_tx = {}
    duplicates_tx = []

    duplicate_value: float = 0

    for index, row in int_df.iterrows():
        tx_id = row[trans_id_col]
        if tx_id in unique_tx:  # duplicate
            duplicates_tx.append(row)
            duplicate_value += row[amount_col]
        else:
            unique_tx[tx_id] = index

    return pd.DataFrame(duplicates_tx), duplicate_value


def write_duplicate_data(
    amount_col_name: str,
    df: pd.DataFrame,
    value: float,
    file_name: str,
    last_row_index=None,
):
    if df.empty:
        return
    number_of_duplicates = len(df)
    amount_column = df.columns.get_loc(amount_col_name)
    if last_row_index is None:
        last_row_index = df.shape[0]
    empty_rows = pd.DataFrame(
        {col: [None] for col in df.columns},
        index=range(last_row_index, last_row_index + 3),
    )
    empty_rows.iat[-2, amount_column] = value
    empty_rows.iat[-1, amount_column] = number_of_duplicates

    df = pd.concat([df, empty_rows], ignore_index=True)

    mode = "a" if os.path.exists(file_name) else "w"
    with pd.ExcelWriter(file_name, engine="openpyxl", mode=mode) as writer:
        sheet_name = "Duplicates"
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def write_missing_ova_data(
    amount_col_name: str,
    df: pd.DataFrame,
    value: float,
    file_name: str,
    last_row_index=None,
):
    if df.empty:
        return
    # If last_row_index is not specified, start from the end of the existing data
    if last_row_index is None:
        last_row_index = df.shape[0]
    number_of_tx = len(df)
    empty_rows = pd.DataFrame(
        {col: [None] for col in df.columns},
        index=range(last_row_index, last_row_index + 3),
    )
    amount_column = df.columns.get_loc(amount_col_name)
    # Concatenate the empty rows with the original DataFrame
    df = pd.concat([df, empty_rows], ignore_index=True)
    # Update the number of tx and value in the new rows
    df.iat[last_row_index + 1, amount_column] = value
    df.iat[last_row_index + 2, amount_column] = number_of_tx
    with pd.ExcelWriter(file_name, engine="openpyxl", mode="a") as writer:
        sheet_name = "Missing OVA Transactions"
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def write_missing_int_data(
    amount_col_name: str,
    df: pd.DataFrame,
    value: float,
    file_name: str,
    last_row_index=None,
):
    if df.empty:
        return
    # If last_row_index is not specified, start from the end of the existing data
    if last_row_index is None:
        last_row_index = df.shape[0]
    number_of_tx = len(df)
    empty_rows = pd.DataFrame(
        {col: [None] for col in df.columns},
        index=range(last_row_index, last_row_index + 3),
    )
    amount_column = df.columns.get_loc(amount_col_name)
    # Concatenate the empty rows with the original DataFrame
    df = pd.concat([df, empty_rows], ignore_index=True)
    # Update the number of tx and value in the new rows
    df.iat[last_row_index + 1, amount_column] = value
    df.iat[last_row_index + 2, amount_column] = number_of_tx
    with pd.ExcelWriter(file_name, engine="openpyxl", mode="a") as writer:
        sheet_name = "Missing INT Transactions"
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def recons_ops(
    file_names: Tuple[str or None, str or None],
    num_lines_of_header: Tuple[int, int],
    alt_recons_name: str,
    file_output_name: str,
    ova_id: str,
    int_id: str,
    alt_ova_id: str,
    alt_int_id: str or None = None,
    *,
    mb_service_name: str or None = None,
    mb_creditDebit_flag: str or None = None,
    mb_status_flag: str or None = None,
    ova_status_flag: str or None = None,
    ova_status_col: str or None = None,
    list_index: int,
):
    if file_names[0] is None and file_names[1] is None:
        return
    service_name_header = ""
    creditDebit_header = ""
    service_name_headers = ["ServiceName", "serviceName", "Service Name"]
    creditDebit_headers = [
        "creditDebitFlag",
        "CreditDebitFlag",
        "DEBITCREDIT",
        "Credit Debit Flag",
    ]
    # -------------------- OVA -------------------
    ova_file_name = file_names[0]  # name of the ova file
    int_file_name = file_names[1]  # name of the integrator file
    ova_header_lines = num_lines_of_header[0]
    int_header_lines = num_lines_of_header[1]
    recons_file = (
        f"{alt_recons_name} - Recons.xlsx"
        if ova_file_name is None
        else f"{ova_file_name[:-5]} - Recons.xlsx"
    )
    ova_file_df: pd.DataFrame or None = None
    int_file_df: pd.DataFrame or None = None

    if ova_file_name is not None:
        # put the data into a dataframe
        ova_file_df = pd.read_excel(
            f"{script_dir}/{ova_file_name}", skiprows=ova_header_lines
        )
        ova_id_name = ova_id
        if ova_status_flag is not None:
            ova_file_df = ova_file_df.loc[
                ova_file_df[ova_status_col] == ova_status_flag
            ]

        for name in creditDebit_headers:
            if name in ova_file_df.columns:
                ova_file_df = ova_file_df[ova_file_df[name] == "C"]
                break
        ova_volume = len(ova_file_df)
        ova_volumes[list_index] = ova_volume
        print(
            f"{file_output_name}_OVA_Volume: {ova_volume}"
        )  # file_output_name is the name that shows for each channel as the script runs
        amount_col = ""
        for name in amount_col_names:
            if name in ova_file_df.columns:
                if not ova_file_df[name].isna().all():
                    amount_col = name
                    break  # check which of the formats the amount column is written in

        ova_value = ova_file_df[amount_col].abs().sum()
        ova_values[list_index] = ova_value
        print(f"{file_output_name} OVA_VALUE : {ova_value}")
        with pd.ExcelWriter(
            recons_file, engine="openpyxl", mode="w"
        ) as writer:  # specify new file name to write to
            ova_file_df.to_excel(
                writer, sheet_name="Sheet1", index=False
            )  # save original data into first sheet of new file

    # ----------------------- INTEGRATOR/ DUPLICATES --------------------
    if int_file_name is not None:
        int_file_df = pd.read_excel(
            f"{script_dir}/{int_file_name}", skiprows=int_header_lines
        )
        int_id_name = int_id
        for name in service_name_headers:
            if name in int_file_df.columns:
                if not int_file_df[name].isna().all():
                    service_name_header = name
                    break
        for name in creditDebit_headers:
            if name in int_file_df.columns:
                if not int_file_df[name].isna().all():
                    creditDebit_header = name
                    break
        if (
            mb_service_name is not None
            and mb_creditDebit_flag is not None
            and mb_status_flag is None
        ):
            int_file_df = int_file_df.loc[
                (int_file_df[service_name_header] == mb_service_name)
                & (int_file_df[creditDebit_header] == mb_creditDebit_flag)
            ]
        elif (
            mb_service_name is not None
            and mb_creditDebit_flag is None
            and mb_status_flag is None
        ):
            int_file_df = int_file_df.loc[
                int_file_df[service_name_header] == mb_service_name
            ]
        if (
            mb_status_flag is not None
            and mb_service_name is None
            and mb_creditDebit_flag is None
        ):
            int_file_df = int_file_df.loc[int_file_df["Status"] == mb_status_flag]
        if int_file_df.empty:
            print(f"No transactions found for {file_output_name}. Confirm")
            return
        int_volume = len(int_file_df)
        int_volumes[list_index] = int_volume
        print(f"{file_output_name}_INT_Volume: {str(int_volume)}")
        amount_col = ""
        for name in amount_col_names:
            if name in int_file_df.columns:
                if not int_file_df[name].isna().all():
                    amount_col = name
                    break
        int_value = int_file_df[amount_col].abs().sum()
        int_values[list_index] = int_value

        print(f"{file_output_name} INT_VALUE : {int_value}")
        dup, dup_val = find_duplicates(int_file_df)
        print(f"Number of duplicates: {len(dup)}")
        print(f"Duplicates value: {dup_val}")
        dup_volumes[list_index] = len(dup)
        dup_values[list_index] = dup_val

        write_duplicate_data(
            amount_col_name=amount_col, df=dup, value=dup_val, file_name=recons_file
        )
    if ova_file_name == f"MPGS{yesterday}.xlsx":
        return
    # ---------------------- MISSING TRANSACTIONS -------------------------
    if ova_file_df is not None and int_file_df is not None:
        ova_id_name = ova_id
        int_id_name = int_id
        ova_file_df[ova_id_name] = ova_file_df[ova_id_name].astype(str)
        int_file_df[int_id_name] = int_file_df[int_id_name].astype(str)

        missing_int_tx = get_missing_tx(
            x=ova_file_df[ova_id_name].astype("string"),
            y=int_file_df[int_id_name].astype("string"),
            alt_x=ova_file_df[alt_ova_id].astype("string"),
            alt_y=int_file_df[alt_int_id].astype("string"),
        ).values

        missing_ova_tx = get_missing_tx(
            x=int_file_df[int_id_name].astype("string"),
            y=ova_file_df[ova_id_name].astype("string"),
            alt_x=int_file_df[alt_int_id].astype("string"),
            alt_y=ova_file_df[alt_ova_id].astype("string"),
        ).values

        int_amount_col = ""
        for name in amount_col_names:
            if name in int_file_df.columns:
                if not int_file_df[name].isna().all():
                    int_amount_col = name
                    break

        ova_amount_col = ""
        for name in amount_col_names:
            if name in ova_file_df.columns:
                if not ova_file_df[name].isna().all():
                    ova_amount_col = name
                    break
        ova_file_df[ova_amount_col] = ova_file_df[ova_amount_col].astype("float")
        int_file_df[int_amount_col] = int_file_df[int_amount_col].astype("float")
        missing_ova_amount_name = ova_amount_col
        missing_int_amount_name = int_amount_col

        missing_ova_data = int_file_df[
            int_file_df[int_id_name].astype("string").isin(missing_ova_tx)
            | int_file_df[alt_int_id].astype("string").isin(missing_ova_tx)
        ]
        missing_ova_value = missing_ova_data[missing_int_amount_name].abs().sum()

        missing_int_data = ova_file_df[
            ova_file_df[ova_id_name].astype("string").isin(missing_int_tx)
            | ova_file_df[alt_ova_id].astype("string").isin(missing_int_tx)
        ]

        missing_int_value = missing_int_data[missing_ova_amount_name].abs().sum()

        write_missing_ova_data(
            amount_col_name=int_amount_col,
            df=missing_ova_data,
            file_name=recons_file,
            value=missing_ova_value,
        )
        write_missing_int_data(
            amount_col_name=ova_amount_col,
            df=missing_int_data,
            file_name=recons_file,
            value=missing_int_value,
        )


def gip_custom(
    ova_files: Tuple[str or None, str or None],
    int_file: str or None,
    num_lines_of_header: Tuple[int, int],
    alt_recons_name: str,
    file_output_name: str,
):
    # ------------------------ OVA ------------------------
    recons_file = (
        f"{alt_recons_name} - Recons.xlsx"
        if ova_files[0] is None
        else f"{ova_files[0][:-5]} - Recons.xlsx"
    )
    ova_files = tuple(item for item in ova_files if item is not None)
    if len(ova_files) == 2:
        ova_df1 = pd.read_excel(f"{script_dir}/{ova_files[0]}")
        ova_df2 = pd.read_excel(f"{script_dir}/{ova_files[1]}")

        ova_volume = len(ova_df1) + len(ova_df2)
        ova_volumes[10] = ova_volume
        amount_col = ""
        for name in amount_col_names:
            if name in ova_df1.columns:
                if not ova_df1[name].isna().all():
                    amount_col = name
                    break
        ova_value = ova_df1[amount_col].abs().sum() + ova_df2[amount_col].abs().sum()
        ova_values[10] = ova_value
        print(f"GIP OVA VOLUME : {ova_volume}")
        print(f"GIP OVA VALUE : {ova_value}")

        with pd.ExcelWriter(
            recons_file, engine="openpyxl", mode="w"
        ) as writer:  # specify new file name to write to
            ova_df1.to_excel(
                writer, sheet_name="Sheet1", index=False
            )  # save original data into first sheet of new file
    # ---------------------------- INT -------------------------
    if int_file is not None:
        int_df = pd.read_excel(f"{script_dir}/{int_file}")
        int_volume = len(int_df)
        int_volumes[10] = int_volume
        amount_col = ""
        for name in amount_col_names:
            if name in int_df.columns:
                if not int_df[name].isna().all():
                    amount_col = name
                    break

        int_value = int_df[amount_col].abs().sum()
        int_values[10] = int_value
        print(f"GIP INT VOLUME {int_volume}")
        print(f"GIP INT VALUE {int_value}")
        dup, dup_val = find_duplicates(int_df)
        write_duplicate_data(
            amount_col_name=amount_col, df=dup, value=dup_val, file_name=recons_file
        )


def get_missing_tx(
    x: pd.Series,
    y: pd.Series,
    alt_x: pd.Series,
    alt_y: pd.Series,
) -> pd.Series:
    x = remove_leading_zeros(x)
    y = remove_leading_zeros(y)

    missing = (~x.astype(str).str.lower().isin(y.astype(str).str.lower())) & (
        ~alt_x.astype(str).str.lower().isin(alt_y.astype(str).str.lower())
    )
    x_missing = x[missing].combine_first(alt_x[missing])
    x_missing[(x_missing == "nan")] = alt_x[missing]
    return x_missing


def remove_leading_zeros(series):
    # Remove leading zeros from the series
    return series.str.replace(r"^0+", "", regex=True)


def reconsLoop():
    recons_ops(
        (
            check_for_file(f"MIGS 01{yesterday}.xlsx"),
            check_for_file(f"MIGS 01 Metabase{yesterday}.xlsx"),
        ),
        num_lines_of_header=(3, 0),
        alt_recons_name=f"MIGS 01{yesterday}",
        file_output_name="MIGS_01",
        list_index=0,
        ova_id="Merchant Transaction Reference",
        int_id="External Payment Request â†’ Institution Trans ID",
        alt_int_id="Institution Trans ID",
        alt_ova_id="Transaction ID",
    )

    recons_ops(
        (
            check_for_file(f"MTN Prompt{yesterday}.xlsx"),
            check_for_file(f"Metabase{yesterday}.xlsx"),
        ),
        num_lines_of_header=(0, 0),
        mb_service_name="MTN Money MADAPI",
        mb_creditDebit_flag="C",
        alt_recons_name=f"MTN Prompt{yesterday}",
        file_output_name="MTN Prompt",
        list_index=1,
        ova_id="External Transaction Id",
        int_id="IntegratorTransId",
        alt_int_id="BillerTransId",
        alt_ova_id="Id",
    )

    recons_ops(
        (
            check_for_file(f"MTN Cashout{yesterday}.xlsx"),
            check_for_file(f"Metabase{yesterday}.xlsx"),
        ),
        num_lines_of_header=(0, 0),
        alt_recons_name=f"MTN Cashout{yesterday}",
        file_output_name="MTN_PORTAL",
        mb_service_name="MTN Money MADAPI",
        mb_creditDebit_flag="D",
        list_index=3,
        ova_id="External Transaction Id",
        int_id="IntegratorTransId",
        alt_int_id="BillerTransId",
        alt_ova_id="Id",
    )
    # recons_ops(
    #     (
    #         check_for_file(f"AirtelTigo Cashout{yesterday}.xlsx"),
    #         check_for_file(f"Metabase{yesterday}.xlsx"),
    #     ),
    #     num_lines_of_header=(4, 0),
    #     alt_recons_name=f"AirtelTigo Cashout{yesterday}",
    #     file_output_name="AIRTEL_CASHOUT",
    #     mb_service_name="AirtelMoney_Slydepay",
    #     mb_creditDebit_flag="D",
    #     list_index=5,
    #     ova_id="Transaction Id",
    #     int_id="IntegratorTransId",
    #     alt_int_id="integratorTransId",
    #     alt_ova_id="Transaction Id",
    # )
    recons_ops(
        (
            check_for_file(f"Vodafone Cashin{yesterday}.xlsx"),
            check_for_file(f"Metabase{yesterday}.xlsx"),
        ),
        num_lines_of_header=(5, 0),
        alt_recons_name=f"Vodafone Cashin{yesterday}",
        file_output_name="VODA CASHIN",
        mb_service_name="Vodafone Cash",
        mb_creditDebit_flag="C",
        list_index=6,
        ova_id="TransId",
        int_id="IntegratorTransId",
        alt_int_id="BillerTransId",
        alt_ova_id="Receipt No.",
    )
    recons_ops(
        (
            check_for_file(f"Vodafone Cashout{yesterday}.xlsx"),
            check_for_file(f"Metabase{yesterday}.xlsx"),
        ),
        num_lines_of_header=(5, 0),
        alt_recons_name=f"Vodafone Cashout{yesterday}",
        file_output_name="VODA CASHOUT",
        mb_creditDebit_flag="D",
        mb_service_name="Vodafone Cash",
        list_index=7,
        ova_id="TransId",
        int_id="IntegratorTransId",
        alt_int_id="BillerTransId",
        alt_ova_id="Receipt No.",
    )
    recons_ops(
        (
            check_for_file(f"Stanbic FI Credit{yesterday}.xlsx"),
            check_for_file(f"Stanbic FI Credit Metabase{yesterday}.xlsx"),
        ),
        num_lines_of_header=(0, 0),
        alt_recons_name=f"Stanbic FI Credit{yesterday}",
        file_output_name="Stanbic FI CREDIT",
        mb_status_flag="CONFIRMED",
        list_index=8,
        ova_id="REMARKS2",
        int_id="External Payment Request → Institution Trans ID",
        alt_int_id="External Payment Request → Institution Trans ID",
        alt_ova_id="REMARKS2",
    )
    gip_custom(
        ova_files=(
            check_for_file(f"slydepay_sending_{GIPdate}'.xlsx"),
            check_for_file(f"slydepay_sendingGhlink_{GIPdate}.xlsx"),
        ),
        num_lines_of_header=(0, 0),
        alt_recons_name=f"slydepay_sending_{yesterday}",
        file_output_name="GIP",
        int_file=check_for_file(f"GIP Metabase{yesterday}.xlsx"),
    )

    recons_ops(
        (
            check_for_file(f"MIGS08{yesterday}.xlsx"),
            check_for_file(f"MiGS_trn{yesterday}.xlsx"),
        ),
        num_lines_of_header=(3, 0),
        alt_recons_name=f"MIGS 08{yesterday}",
        file_output_name="BB MIG",
        ova_id="Transaction Number",
        int_id="Receipt No.",
        alt_ova_id="Order ID",
        alt_int_id="Receipt No.",
        mb_status_flag="CONFIRMED",
        list_index=21,
    )
    try:
        recons_ops(
            (
                check_for_file(f"MPGS{yesterday}.xlsx"),
                check_for_file(f"MPGS_trn{yesterday}.xlsx"),
            ),
            num_lines_of_header=(0, 0),
            alt_recons_name=f"MPGS{yesterday}",
            file_output_name="MPGS",
            mb_status_flag="CONFIRMED",
            alt_int_id="BillerTransId",
            alt_ova_id="Acquirer Transaction ID",
            list_index=22,
            ova_id="Acquirer Transaction ID",
            int_id="BillerTransId",
        )
    except:
        recons_ops(
            (
                check_for_file(f"MPGS{yesterday}.xlsx"),
                check_for_file(f"MPGS_trn{yesterday}.xlsx"),
            ),
            num_lines_of_header=(0, 0),
            alt_recons_name=f"MPGS{yesterday}",
            file_output_name="MPGS",
            mb_status_flag="CONFIRMED",
            alt_int_id="Bill Er Trans ID",
            alt_ova_id="Acquirer Transaction ID",
            list_index=22,
            ova_id="Acquirer Transaction ID",
            int_id="Bill Er Trans ID",
        )
    recons_ops(
        (
            check_for_file(f"Quipu{yesterday}.xlsx"),
            check_for_file(f"Quipu KC{yesterday}.xlsx"),
        ),
        num_lines_of_header=(0, 0),
        alt_recons_name=f"Quipu{yesterday}",
        file_output_name="QUIPU",
        ova_status_flag="SUCCESS",
        ova_status_col="Status",
        list_index=23,
        ova_id="Order Code",
        int_id="Universal Transaction Reference",
        alt_int_id="Receipt No",
        alt_ova_id="Order Code",
    )
    try:
        recons_ops(
            (
                check_for_file(f"KR MTN Credit{yesterday}.xlsx"),
                check_for_file(f"KR MTN Disb_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(0, 0),
            alt_recons_name=f"KR MTN Credit{yesterday}",
            file_output_name="MTN_KR_Credit",
            list_index=24,
            ova_id="External Transaction Id",
            int_id="Integrator Trans ID",
            alt_int_id="Bill Er Trans ID",
            alt_ova_id="Id",
        )
    except:
        recons_ops(
            (
                check_for_file(f"KR MTN Credit{yesterday}.xlsx"),
                check_for_file(f"KR MTN Disb_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(0, 0),
            alt_recons_name=f"KR MTN Credit{yesterday}",
            file_output_name="MTN_KR_Credit",
            list_index=24,
            ova_id="External Transaction Id",
            int_id="IntegratorTransId",
            alt_int_id="BillerTransId",
            alt_ova_id="Id",
        )
    try:
        recons_ops(
            (
                check_for_file(f"KR MTN Debit{yesterday}.xlsx"),
                check_for_file(f"KR MTN Coll_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(0, 0),
            alt_recons_name=f"KR MTN Debit{yesterday}",
            file_output_name="MTN_KR_Debit",
            list_index=25,
            ova_id="External Transaction Id",
            int_id="Integrator Trans ID",
            alt_ova_id="Id",
            alt_int_id="Bill Er Trans ID",
        )
    except:
        recons_ops(
            (
                check_for_file(f"KR MTN Debit{yesterday}.xlsx"),
                check_for_file(f"KR MTN Coll_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(0, 0),
            alt_recons_name=f"KR MTN Debit{yesterday}",
            file_output_name="MTN_KR_Debit",
            list_index=25,
            ova_id="External Transaction Id",
            int_id="IntegratorTransId",
            alt_ova_id="Id",
            alt_int_id="BillerTransId",
        )
    # recons_ops(
    #     (
    #         check_for_file(f"KR AirtelTigo{yesterday}.xlsx"),
    #         check_for_file(f"KR AirtelTigo Coll_mBase{yesterday}.xlsx"),
    #     ),
    #     num_lines_of_header=(0, 0),
    #     alt_recons_name=f"AirtelTigo Cashin{yesterday}",
    #     file_output_name="AIRTEL_KR_Cashin",
    #     ova_status_flag="Merchant Payment",
    #     ova_status_col="Service Type",
    #     mb_status_flag="CONFIRMED",
    #     list_index=26,
    #     ova_id="External Transaction Id",
    #     int_id="Transaction Id",
    #     alt_int_id="",
    #     alt_ova_id="",
    # )
    recons_ops(
        (
            check_for_file(f"KR AirtelTigo{yesterday}.xlsx"),
            check_for_file(f"KR AirtelTigo Disb_mBase{yesterday}.xlsx"),
        ),
        num_lines_of_header=(0, 0),
        alt_recons_name=f"AirtelTigo Cashout{yesterday}",
        file_output_name="AIRTEL_KR_Cashout",
        ova_status_flag="Cash in",
        ova_status_col="Service Type",
        mb_status_flag="CONFIRMED",
        list_index=27,
        ova_id="External Transaction Id",
        int_id="Transaction Id",
        alt_int_id="Transaction Id",
        alt_ova_id="External Transaction Id",
    )
    try:
        recons_ops(
            (
                check_for_file(f"KR Vodafone Cashin{yesterday}.xlsx"),
                check_for_file(f"KR Vodafone Coll_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(5, 0),
            alt_recons_name=f"KR VODA Cashin{yesterday}",
            file_output_name="Voda KR Cashin",
            mb_status_flag="CONFIRMED",
            list_index=28,
            ova_id="TransId",
            int_id="Integrator Trans ID",
            alt_int_id="Bill Er Trans ID",
            alt_ova_id="Receipt No.",
        )
    except:
        recons_ops(
            (
                check_for_file(f"KR Vodafone Cashin{yesterday}.xlsx"),
                check_for_file(f"KR Vodafone Coll_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(5, 0),
            alt_recons_name=f"KR VODA Cashin{yesterday}",
            file_output_name="Voda KR Cashin",
            mb_status_flag="CONFIRMED",
            list_index=28,
            ova_id="TransId",
            int_id="Transaction Id",
            alt_int_id="Receipt No.",
            alt_ova_id="Receipt No.",
        )

    try:
        recons_ops(
            (
                check_for_file(f"KR Vodafone Cashout{yesterday}.xlsx"),
                check_for_file(f"KR Vodafone Disb_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(5, 0),
            alt_recons_name=f"KR Voda Cashout{yesterday}",
            file_output_name="Voda KR Cashout",
            mb_status_flag="CONFIRMED",
            list_index=29,
            ova_id="TransId",
            int_id="Integrator Trans ID",
            alt_int_id="Bill Er Trans ID",
            alt_ova_id="Receipt No.",
        )
    except:
        recons_ops(
            (
                check_for_file(f"KR Vodafone Cashout{yesterday}.xlsx"),
                check_for_file(f"KR Vodafone Disb_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(5, 0),
            alt_recons_name=f"KR Voda Cashout{yesterday}",
            file_output_name="Voda KR Cashout",
            mb_status_flag="CONFIRMED",
            list_index=29,
            ova_id="TransId",
            int_id="Transaction Id",
            alt_int_id="Receipt No.",
            alt_ova_id="Receipt No.",
        )
    print("Completed")


def main():
    root = ctk.CTk()
    root.title("Kowri Recons")
    root.option_add("*tearOff", False)
    root._set_appearance_mode("light")
    root.geometry("500x550")

    def render_recons_frame(tab):
        recons_frame = ctk.CTkFrame(tab, fg_color="white")

        logo_img = ctk.CTkImage(
            Image.open(resource_path(f"{script_dir}/assets/KowriLogo.png")),
            size=(200, 70),
        )
        label = ctk.CTkLabel(recons_frame, image=logo_img, text="", fg_color="white")
        label.pack()

        yesterday_frame = ttk.LabelFrame(
            recons_frame, text="Set Recons Date", width=500, height=400
        )
        yesterday_frame.pack()
        yesterday_frame.pack_propagate(False)

        yesterady_msg = ctk.CTkLabel(
            yesterday_frame, text="Recons for yesterday?", text_color="green"
        )
        yesterady_msg.pack()

        recons_var = ctk.StringVar()
        recons_var.set("yes")

        def yesterdayRecons():
            dayCombo.configure(state="disabled")
            dayCombo.set((date.today().day) - 1)
            dayLabel.configure(text_color="grey")

            monthCombo.configure(state="disabled")
            monthCombo.set(date.today().month)
            monthLabel.configure(text_color="grey")

            yearCombo.configure(state="disabled")
            yearCombo.set(date.today().year)
            yearLabel.configure(text_color="grey")

            today = datetime.today().strftime("%d_%b_%y")

        def customDateRecons():
            dayCombo.configure(state="enabled")
            dayCombo.current(0)
            dayLabel.configure(text_color="green")

            monthCombo.configure(state="enabled")
            monthCombo.current(0)
            monthLabel.configure(text_color="green")

            yearCombo.configure(state="enabled")
            yearLabel.configure(text_color="green")

        def topLevelConsole():
            consoleWindow = ctk.CTkToplevel()
            consoleWindow.title("Results")
            consoleWindow.geometry("500x650")
            consoleWindow.attributes("-topmost", True)

            consoleWindow.columnconfigure(index=0, weight=1)
            consoleWindow.columnconfigure(index=1, weight=1)
            consoleWindow.columnconfigure(index=2, weight=1)
            consoleWindow.rowconfigure(index=0, weight=1)
            consoleWindow.rowconfigure(index=1, weight=1)
            consoleWindow.rowconfigure(index=2, weight=1)

            console_frame = ttk.Frame(consoleWindow)
            console_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

            console = Console(console_frame)
            console.pack(fill=tk.BOTH, expand=True)

            try:
                run_recons()
                reconsLoop()
            finally:
                progress_bar.stop()
                progress_bar.pack_forget()
            # consoleWindow.mainloop()

        def loadingProgress():
            root.update()
            progress_bar.pack()
            progress_bar.start()

        def startProgress():
            threading.Thread(target=loadingProgress).start()
            # progressThread.start()

        def startWindow():
            threading.Thread(target=topLevelConsole).start()
            # consoleThread.start()

        def run_recons():
            global yesterday
            global recons_yesterday
            global GIPdate
            global current_month
            global ova_values
            global ova_volumes
            global int_volumes
            global int_values
            global dup_volumes
            global dup_values
            global list_index

            if recons_var.get() == "no":
                customDay = dayCombo.get()
                customMonth = monthCombo.get()
                customYear = yearCombo.get()[-2:]
                GIPday = int(customDay) + 1
                if int(customDay) < 10:
                    customDay = "0" + customDay
                else:
                    customDay = customDay
                month_name = calendar.month_abbr[int(customMonth)]
                if int(customMonth) < 10:
                    customMonth = "0" + customMonth
                else:
                    customMonth = customMonth
                yesterday = "_" + customDay + " " + month_name + "_" + customYear
                print(yesterday)

                GIPmonth = 0
                GIPday = 0

                recons_yesterday = (
                    "20"
                    + customYear
                    + "-"
                    + customMonth
                    + "-"
                    + customDay
                    + " 00:00:00"
                )
                print(recons_yesterday)

                if int(customMonth) < 10:
                    GIPmonth = customMonth

                if int(GIPday) < 10:
                    GIPday = "0" + str(GIPday)
                GIPdate = str("20" + str(customYear) + str(GIPmonth) + str((GIPday)))
                print(GIPdate)

            else:
                # Yesterday's date
                yesterday = (datetime.now() - timedelta(1)).strftime("_%d %b_%y")
                print(yesterday)

                # Current month
                current_month = calendar.month_abbr[date.today().month].upper()
                print(current_month)

                recons_yesterday = str(
                    (datetime.now() - timedelta(1)).strftime("%Y-%m-%d") + " 00:00:00"
                )
                print(recons_yesterday)

                if date.today().month < 10:
                    GIPmonth = "0" + str(date.today().month)
                else:
                    GIPmonth = str(date.today().month)

                if date.today().day < 10:
                    GIPday = "0" + str(date.today().day + 1)
                else:
                    GIPday = str(date.today().day + 1)
                GIPdate = str(str(date.today().year) + str(GIPmonth) + str((GIPday)))
                print(GIPdate)

            ova_volumes = [0] * 30
            ova_values = [0] * 30
            int_volumes = [0] * 30
            int_values = [0] * 30
            dup_volumes = [0] * 30
            dup_values = [0.00] * 30
            list_index = 0

        day_value_list = []
        for i in range(1, 32):
            day_value_list.append(i)

        month_value_list = []
        for i in range(1, 13):
            month_value_list.append(i)

        yesRadio = ttk.Radiobutton(
            yesterday_frame,
            text="Yes",
            variable=recons_var,
            value="yes",
            command=yesterdayRecons,
        )
        yesRadio.pack(anchor=ctk.W)
        noRadio = ttk.Radiobutton(
            yesterday_frame,
            text="No",
            variable=recons_var,
            value="no",
            command=customDateRecons,
        )
        noRadio.pack(anchor=ctk.W)

        dayLabel = ctk.CTkLabel(yesterday_frame, text="Day", text_color="grey")
        dayLabel.pack()
        dayCombo = ttk.Combobox(
            yesterday_frame, state="disabled", values=day_value_list
        )
        dayCombo.current(0)
        dayCombo.set((date.today().day) - 1)
        dayCombo.pack()

        monthLabel = ctk.CTkLabel(yesterday_frame, text="Month", text_color="grey")
        monthLabel.pack()
        monthCombo = ttk.Combobox(
            yesterday_frame, state="disabled", values=month_value_list
        )
        monthCombo.current(0)
        monthCombo.set(date.today().month)
        monthCombo.pack()

        yearLabel = ctk.CTkLabel(yesterday_frame, text="Year", text_color="grey")
        yearLabel.pack()
        yearCombo = ttk.Combobox(
            yesterday_frame, state="disabled", values=["2022", "2023", "2024"]
        )
        yearCombo.current(0)
        yearCombo.set(date.today().year)
        yearCombo.pack()

        startButton = ctk.CTkButton(
            yesterday_frame,
            text="Start",
            fg_color="green",
            hover_color="#1bcf48",
            command=lambda: (startWindow(), startProgress()),
        )
        startButton.pack(pady=(20, 0))

        progress_bar = ttk.Progressbar(yesterday_frame, mode="indeterminate")
        progress_bar.pack_forget()
        recons_frame.pack(fill="both", expand=True)

    # Make the window interactive
    root.columnconfigure(index=0, weight=1)
    root.columnconfigure(index=1, weight=1)
    root.columnconfigure(index=2, weight=1)
    root.rowconfigure(index=0, weight=1)
    root.rowconfigure(index=1, weight=1)
    root.rowconfigure(index=2, weight=1)

    style = ttk.Style(root)
    # Import the tcl file
    root.tk.call(
        "source",
        resource_path(f"{script_dir}/assets/Forest-ttk-theme-master/forest-light.tcl"),
    )

    # Set the theme with the theme_use method
    style.theme_use("forest-light")

    tab_control = ttk.Notebook(root)
    recons_tab = ttk.Frame(tab_control)
    logs_tab = ttk.Frame(tab_control)

    tab_control.add(recons_tab, text="Recons")
    tab_control.add(logs_tab, text="Logs")

    render_recons_frame(recons_tab)

    tab_control.pack(side=ctk.LEFT, fill="both", expand=True)

    root.mainloop()


if __name__ == "__main__":
    main()
