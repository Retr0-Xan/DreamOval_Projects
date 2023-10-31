# version 1.1.0
from openpyxl import load_workbook, worksheet
from openpyxl.worksheet import worksheet
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
import pandas as pd
from typing import Literal, Tuple

# from dateTime import yesterday, GIPdate, current_month, recons_yesterday
import os
from datetime import datetime
from datetime import date
from datetime import timedelta


def get_date(
    date: date, format: Literal["gip"] | Literal["normal"] | Literal["recons"]
):
    if format == "gip":
        return date.strftime("%Y%m%d")  # Output date in format: 20231201
    elif format == "recons":
        return date.strftime("%Y-%m-%d")
    else:
        return date.strftime("%-d %b_%y")  # Output date in format: 1_Jan_23


prompt = input("Recons for yesterday? (Y/N) :")
if prompt.upper() == "Y":
    date_ = date.today() - timedelta(1)
else:
    print("Please enter the date...")
    day = int(input("Recons Day (1-31): "))
    month = int(input("Recons Month (1-12): "))
    year = int(input("Recons Year (eg 2023): "))
    date_ = datetime(year=year, month=month, day=day)

yesterday = f"_{get_date(date_, format='normal')}"
print(yesterday)

recons_yesterday = f"{get_date(date_, format='recons')} 00:00:00"
print(recons_yesterday)

current_month = date_.strftime("%b").upper()  # JAN, FEB
print(current_month)

GIPdate = get_date(date=date_ + timedelta(1), format="gip")
print(GIPdate)


def get_column(
    keyword: str,
    sheet: worksheet.Worksheet,
    first_row: int,
    last_row: int,
    first_col: int,
    last_col: int,
):
    for i in range(first_row, last_row + 1):
        for j in range(first_col, last_col + 1):
            if sheet[str(get_column_letter(j)) + str(i)].value == keyword:
                return get_column_letter(j)


dir_list = os.listdir()


def metabase_file_exists() -> bool:
    dir_list = os.listdir()
    # Check if metabase file is available
    metabase_file_name = "Metabase" + yesterday + ".xlsx"
    for file in dir_list:
        if file == metabase_file_name:
            return True
    return False


meta_file_found = metabase_file_exists()


def find_duplicates(
    file_name: str,
    skip_rows: int = 0
):
    tx_id_col_names = ["integratorTransId", "IntegratorTransId"]
    amount_col_names = ["Amount", "amount"]

    int_df = pd.read_excel(file_name, skiprows=skip_rows)
    int_df = int_df.loc[
        (int_df["ServiceName"] == "MTN OVA") & (int_df["CreditDebitFlag"] == "C")
    ]

    trans_id_col = ""
    for name in tx_id_col_names:
        if name in int_df.columns:
            trans_id_col = name
            break
        
    amount_col = ""
    for name in amount_col_names:
        if name in int_df.columns:
            amount_col = name
            break

    if trans_id_col == "":
        raise Exception("No transaction id column found")

    unique_tx = {}
    duplicates_tx = []

    duplicate_value: float = 0

    for index, row in int_df.iterrows():
        tx_id = row[trans_id_col]
        if tx_id in unique_tx:  # duplicate
            duplicates_tx.append(row)
            duplicate_value += row[amount_col]
        else:
            unique_tx[tx_id] = index

    with pd.ExcelWriter(file_name, engine="openpyxl", mode="a") as writer:
        sheet_name = "Duplicates"

        dup_df = pd.DataFrame(duplicates_tx)
        dup_df.to_excel(writer, sheet_name=sheet_name, index=False)

def run_recons(file_names: Tuple[str, str], num_lines_of_header: Tuple[int, int]):
    ova_file_name = file_names[0]
    int_file_name = file_names[1]
    ova_header_lines = num_lines_of_header[0]
    int_header_lines = num_lines_of_header[1]

    file_found = False
    match_ova_found = False
    for file in dir_list:
        if file == ova_file_name:
            file_found = True
            match_ova_found = True
            wb = load_workbook(ova_file_name, read_only=True)
            sheet = wb[wb.sheetnames[0]]
            wb.active = wb[wb.sheetnames[0]]

            # DEFINE MAX AND MIN COLUMNS AND ROWS
            first_col: int = wb.active.min_column
            last_col: int = wb.active.max_column
            first_row: int = wb.active.min_row
            last_row: int = wb.active.max_row

            OVA_VOLUME1 = (last_row + 1) - (first_row + ova_header_lines)
            print("MIGS_01_OVA_Volume: " + str(OVA_VOLUME1))

            amount_total = 0.00

            amount_column = get_column(
                "Amount", sheet, first_row, last_row, first_col, last_col
            )  # TODO: Some have different names instead of "Amount"
            if (
                amount_column == None
            ):  # TODO: Might need to handle this differently, other than just return
                return

            for i in range(first_row + 4, last_row + 1):
                amount_total += abs(float(sheet[amount_column + str(i)].value))

            print("MIGS_01_OVA_Sum: " + str(amount_total))
            OVA_VALUE1 = amount_total
            ova_id_col = get_column(
                "Merchant Transaction Reference",
                sheet,
                first_row,
                last_row,
                first_col,
                last_col,
            )
            tmp_id_col = ova_id_col

            alternate_id_col = get_column(
                "Transaction ID", sheet, first_row, last_row, first_col, last_col
            )
            wb.close()

            match_ova = ova_file_name.rpartition(".")[0] + " - Recons.xlsx"
            wb.save(match_ova)

            wb = load_workbook(match_ova)
            ws1 = wb.create_sheet("Duplicates")
            ws1.title = "Duplicates"

            ws1 = wb.create_sheet("Missing Integrator Transactions")
            ws1.title = "Missing Integrator Transactions"

            ws1 = wb.create_sheet("Missing OVA Transactions")
            ws1.title = "Missing OVA Transactions"

            wb.close()
            wb.save(match_ova)
            break
        else:
            file_found = False

    if file_found is False:
        OVA_VOLUME1 = 0
        OVA_VALUE1 = 0

    #
    #
    #
    file_found = False
    match_int_found = False
    #  ##################################################### SLYDEPAY01 INT #################################################
    for file in dir_list:
        if file == int_file_name:
            file_found = True
            match_int_found = True
            wb = load_workbook(int_file_name)
            sheet = wb[wb.sheetnames[0]]
            wb.active = wb[wb.sheetnames[0]]
            file = pd.read_excel(int_file_name)
            match_INT = int_file_name

            # DEFINE MAX AND MIN COLUMNS AND ROWS
            first_col = wb.active.min_column
            last_col = wb.active.max_column
            first_row = wb.active.min_row
            last_row = wb.active.max_row

            MIGS_01_INT_Volume = (last_row + 1) - (first_row + 1)
            INT_VOLUME1 = MIGS_01_INT_Volume
            print("MIGS_01_INT_Volume: " + str(MIGS_01_INT_Volume))

            MIGS_01_INT_Sum = 0.00

            int_id_col = get_column(
                "External Payment Request â†’ Institution Trans ID",
                sheet,
                first_row,
                last_row,
                first_col,
                last_col,
            )
            id_Col = "External Payment Request â†’ Institution Trans ID"
            if int_id_col is None:
                int_id_col = get_column(
                    "Institution Trans ID",
                    sheet,
                    first_row,
                    last_row,
                    first_col,
                    last_col,
                )
                id_Col = "Institution Trans ID"
            amount_column = get_column(
                "Amount", sheet, first_row, last_row, first_col, last_col
            )
            if amount_column is None:
                amount_column = get_column(
                    "Amount ($)", sheet, first_row, last_row, first_col, last_col
                )
            if amount_column is None:
                amount_column = get_column(
                    "Actual Amount ($)", sheet, first_row, last_row, first_col, last_col
                )
            for i in range(first_row + 1, last_row + 1):
                MIGS_01_INT_Sum = MIGS_01_INT_Sum + abs(
                    float(sheet[amount_column + str(i)].value)
                )

            print("MIGS_01_INT_Sum: " + str(MIGS_01_INT_Sum))
            INT_VALUE1 = MIGS_01_INT_Sum
            wb.close()
            wb.save(int_file_name)
            
            # ---------------------------------------- Get duplicates -----------------------------------------------------
            find_duplicates(match_ova, int_file_name, first_row, last_row, int_id_col, id_Col, file)


            # ---------------------------------------- MISSING INT TRANSACTIONS -----------------------------------------
            if match_ova_found is True and match_int_found is True:
                Owb = load_workbook(match_ova)
                Osheet = Owb[Owb.sheetnames[0]]
                Owb.active = Owb[Owb.sheetnames[0]]
                file = pd.read_excel(match_ova, header=[3])

                Iwb = load_workbook(match_INT)
                Isheet = Iwb[Iwb.sheetnames[0]]
                Iwb.active = Iwb[Iwb.sheetnames[0]]

                Ofirst_col = Owb.active.min_column
                Olast_col = Owb.active.max_column
                Ofirst_row = Owb.active.min_row
                Olast_row = Owb.active.max_row

                Ifirst_col = Iwb.active.min_column
                Ilast_col = Iwb.active.max_column
                Ifirst_row = Iwb.active.min_row
                Ilast_row = Iwb.active.max_row

                missing_INT_list = []
                missing_OVA_list = []
                headerChecker = True
                matchFound = False
                counter = 0

                for i in range(Ofirst_row + 4, Olast_row + 1):
                    for j in range(Ifirst_row + 1, Ilast_row + 1):
                        if (
                            Osheet[ova_id_col + str(i)].value is None
                            or Osheet[ova_id_col + str(i)].value == "#VALUE!"
                        ):
                            ova_id_col = alternate_id_col
                            matchFound = False
                            break
                        if str(Osheet[ova_id_col + str(i)].value) == str(
                            Isheet[int_id_col + str(j)].value
                        ):
                            matchFound = True
                            break
                        else:
                            matchFound = False
                    if matchFound is False:
                        missing_INT_list.append(Osheet[ova_id_col + str(i)].value)
                        ova_id_col = tmp_id_col
                print(f"Missing integrator transactions: {missing_INT_list}")

                count = 0
                for transaction in missing_INT_list:
                    data = file[file["Merchant Transaction Reference"] == transaction]
                    if data.empty:
                        data = file[file["Transaction ID"] == transaction]
                    if counter > 0:
                        headerChecker = False
                    with pd.ExcelWriter(
                        match_ova,
                        mode="a",
                        engine="openpyxl",
                        if_sheet_exists="overlay",
                    ) as writer:
                        data.to_excel(
                            writer,
                            sheet_name="Missing Integrator Transactions",
                            startrow=count,
                            header=headerChecker,
                        )
                        if headerChecker is False:
                            count += 1
                        else:
                            count += 2
                        counter += 1

                Iwb.close()
                Owb.close()

                # ----------------------------------------- MISSING OVA TRANSACTIONS -------------------------------------------
                Owb = load_workbook(match_ova)
                Osheet = Owb[Owb.sheetnames[0]]
                Owb.active = Owb[Owb.sheetnames[0]]

                Iwb = load_workbook(match_INT)
                Isheet = Iwb[Iwb.sheetnames[0]]
                Iwb.active = Iwb[Iwb.sheetnames[0]]
                file = pd.read_excel(match_INT)

                Ofirst_col = Owb.active.min_column
                Olast_col = Owb.active.max_column
                Ofirst_row = Owb.active.min_row
                Olast_row = Owb.active.max_row

                Ifirst_col = Iwb.active.min_column
                Ilast_col = Iwb.active.max_column
                Ifirst_row = Iwb.active.min_row
                Ilast_row = Iwb.active.max_row

                missing_INT_list = []
                missing_OVA_list = []
                matchFound = False
                headerChecker = True
                counter = 0

                for i in range(Ifirst_row + 1, Ilast_row + 1):
                    for j in range(Ofirst_row + 1, Olast_row + 1):
                        if str(Isheet[int_id_col + str(i)].value) == str(
                            Osheet[ova_id_col + str(j)].value
                        ):
                            matchFound = True
                            break
                        else:
                            matchFound = False
                    if matchFound is False:
                        missing_OVA_list.append(Isheet[int_id_col + str(i)].value)
                print(f"Missing OVA transactions: {missing_OVA_list}")
                count = 0
                for transaction in missing_OVA_list:
                    data = file[file[id_Col] == transaction]
                    if counter > 0:
                        headerChecker = False
                    with pd.ExcelWriter(
                        match_ova,
                        mode="a",
                        engine="openpyxl",
                        if_sheet_exists="overlay",
                    ) as writer:
                        data.to_excel(
                            writer,
                            sheet_name="Missing OVA Transactions",
                            startrow=count,
                            header=headerChecker,
                        )
                        if headerChecker is False:
                            count += 1
                        else:
                            count += 2
                        counter += 1
                Iwb.close()
                Owb.close()
            break
        else:
            file_found = False

    if file_found is False:
        INT_VOLUME1 = 0
        INT_VALUE1 = 0
        DUPLICATES_VOLUME1 = 0
        DUPLICATES_VALUE1 = 0
