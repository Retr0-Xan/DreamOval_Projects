import pandas as pd
from pandas import *
from openpyxl import load_workbook
from datetime import date
import os

# setting the path for the settlement advice file and putting it into a dataframe
settlement_advice_file = "./settlementAuto/Data/settlement_advice.csv"
settle_adv_df = pd.read_csv(settlement_advice_file)


# the bulk generator has some empty rows which affect the dataframe. the empty rows actually contain the formula to get the instution code.
# This function clears those rows
def clear_rows_with_value(sheet, column):
    for row in range(sheet.max_row, 1, -1):
        cell_value = sheet.cell(row=row, column=column).value
        if (
            cell_value
            == f"=_xlfn.VALUETOTEXT(IF(ISNA(VLOOKUP(A{row},'Institutional Codes'!$A$1:$B$24,2,FALSE)),\"\",VLOOKUP(A{row},'Institutional Codes'!$A$1:$B$24,2,FALSE)))"
        ):
            sheet.delete_rows(row)


# loading the generator into openpyxl. Openpyxl because we use openpyxl to populate the sheet
wb = load_workbook("settlementAuto/kowri_payment_file_generator.xlsm")
sheet = wb[wb.sheetnames[0]]
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

# codes dictionary.To add a code, specify the code as key(on the left) and its corresponding bank as value(on the right)
codes = {
    "300303": "Absa Bank",
    "300302": "Standard Chartered Bank",
    "300304": "GCB Bank",
    "300305": "National Investment Bank",
    "300307": "Agricultural Development Bank",
    "300309": "Universal Merchant Bank",
    "300310": "Republic Bank",
    "300311": "Zenith Bank Limited",
    "300312": "Ecobank Ghana",
    "300313": "CAL Bank",
    "300317": "Prudential Bank Limited",
    "300318": "Stanbic Bank",
    "300322": "Guaranty Trust (GH) Limited",
    "300325": "United Bank of Africa",
    "300329": "Access Bank Limited",
    "300334": "First National Bank",
    "300487": "Kowri",
    "23354": "MTN Mobile Money",
    "23324": "MTN Mobile Money",
    "23355": "MTN Mobile Money",
    "23359": "MTN Mobile Money",
    "23325": "MTN Mobile Money",
    "23356": "AirtelTigo Money",
    "23357": "AirtelTigo Money",
    "23326": "AirtelTigo Money",
    "23327": "AirtelTigo Money",
    "23320": "Vodafone Cash",
    "23350": "Vodafone Cash",
    "300323": "Fidelity Bank",
    "300320": "Bank of Africa",
    "300316": "First Atlantic Bank",
    "300319": "FBN Bank",
}
# institution codes dictionary
institution_codes = {
    "Kowri": "kowri",
    "MTN Mobile Money": "mtn-money",
    "AirtelTigo Money": "airteltigo-money",
    "Vodafone Cash": "vodafone-cash",
    "National Investment Bank": "nib-account-fi-service",
    "Prudential Bank Limited": "prudential-account-fi-service",
    "Guaranty Trust (GH) Limited": "gt-account-fi-service",
    "First National Bank": "First National Bank fnb-account-fi-service",
    "Universal Merchant Bank": "umb-account-fi-service",
    "Zenith Bank Limited": "zenith-account-fi-service",
    "Access Bank Limited": "access-account-fi-service",
    "CAL Bank": "cal-account-fi-service",
    "Standard Chartered Bank": "standardchartered-account-fi-service",
    "Ecobank Ghana": "ecobank-account-fi-service",
    "Absa Bank": "absa-account-fi-service",
    "GCB Bank": "gcb-account-fi-service",
    "Stanbic Bank": "stanbic-account-fi-service",
    "Agricultural Development Bank": "adb-account-fi-service",
    "United Bank of Africa": "uba-account-fi-service",
    "FBN Bank": "fnb-account-fi-service",
    "Fidelity Bank": "fidelity-account-fi-service",
    "First Atlantic Bank": "first-atlantic-account-fi-service",
    "Republic Bank": "republic-account-fi-service",
    "Bank of Africa": "bank-africa-account-fi-service",
}

# list to hold settlements that couldn't be completed with the script
unfinished_settlements = []

# this stores all the account numbers in the settlement advice file
sett_accNum = settle_adv_df["Account Number"].astype(str)
rowNum = 0
useRow = 7
# going through the account numbers to determine their banks and institution codes
for count, (i, acc) in enumerate(enumerate(sett_accNum)):
    rowNum = i + 7

    if acc.startswith("233"):
        useAcc = acc[0:5]
        finAcc = acc
        skiprowBool = False
        useRow += 1

    elif acc.startswith("300"):
        useAcc = acc[0:6]
        finAcc = acc[7:]
        skiprowBool = False
        useRow += 1
    else:
        # if the account number value for the specific row is empty put that merchant into the unfinished settlements list.
        if acc == "nan":
            unfinished_settlements.append(settle_adv_df["Service Name"].iloc[i])
        else:
            unfinished_settlements.append(acc)
        if skiprowBool:
            continue
        else:
            skiprowBool = True
        continue
    # writing the values into the generator file
    if useAcc in codes:
        # print(f"the account {acc} is {codes[useAcc]}")
        sheet[f"B{useRow}"].value = finAcc
        sheet[f"A{useRow}"].value = codes[useAcc]
        sheet[f"C{useRow}"].value = settle_adv_df["Account Name"].iloc[i]
        sheet[f"D{useRow}"].value = settle_adv_df["Settlement Amount"].iloc[i]
        sheet[f"E{useRow}"].value = f"{acc[0:11]}_{date.today()}"
        sheet[
            f"F{useRow}"
        ].value = f"KB_Settlement_{settle_adv_df['Service Name'].iloc[i]}"
        sheet[f"G{useRow}"].value = institution_codes[codes[useAcc]]
    else:
        useRow -= 1
        continue
column_to_check = 7
# saving the unfinshed settlements as a csv file. Find this file in Data folder with name "Manual_settlements_pending.csv"
pd.DataFrame([unfinished_settlements]).to_csv(
    "./settlementAuto/Data/Manual_settlements_pending.csv", index=False
)

#clearing the empty rows of the generator files
clear_rows_with_value(sheet, column_to_check)
wb.close()

#saving it as a temporary xlsm file because openpyxl can't save as csv.
wb.save(f"./settlementAuto/Data/temp.xlsm")

#now we read our temporary file into a dataframe
output_df = pd.read_excel(f"./settlementAuto/Data/temp.xlsm", skiprows=6)

#and save that dataframe as a csv file with pandas
output_df.to_csv(
    f"./settlementAuto/Data/bulk_settlement_{date.today()}.csv", index=False
)
#now we can remove the temporary file
os.remove("./settlementAuto/Data/temp.xlsm")

#list all the unfinished settlements
print(f"Unfinished settlements {unfinished_settlements}")
