from openpyxl import load_workbook
import pandas as pd

Owb = load_workbook('MTN KR Debit_OVA_01_Oct_22.xlsx')
Osheet = Owb['Sheet1']

Iwb = load_workbook('MTN KR Debit INT_01_Oct_22.xlsx')
Isheet = Iwb['MTN KR Debit Metabase_01_Oct_22']

Ofirst_col = Owb.active.min_column
Olast_col = Owb.active.max_column
Ofirst_row = Owb.active.min_row
Olast_row = Owb.active.max_row

Ifirst_col = Iwb.active.min_column
Ilast_col = Iwb.active.max_column
Ifirst_row = Iwb.active.min_row
Ilast_row = Iwb.active.max_row

missing_INT_list = []
missing_OVA_list = []
matchFound = False


def missing_INT_transactions():
    for i in range(Ofirst_row+1, Olast_row+1):
        for j in range(Ifirst_row+1, Ilast_row+1):
            if Osheet['B' + str(i)].value == Isheet['H'+str(j)].value:
                matchFound = True
                break
            else:
                matchFound = False
        if matchFound == False:
            missing_INT_list.append(Osheet['B' + str(i)].value)
    print(f"Missing integrator transactions: {missing_INT_list}")
    wb = load_workbook('test_file.xlsx')
    ws2 = wb.create_sheet('Missing Integrator Transactions')
    ws2.title = 'Missing Integrator Transactions'
    file = pd.read_excel('MTN KR Debit_OVA_01_Oct_22.xlsx')
    count = 0
    for transaction in missing_INT_list:
        data = file[file['External Transaction Id'] == transaction]
        with pd.ExcelWriter('test_file.xlsx', mode='a', engine="openpyxl", if_sheet_exists='overlay') as writer:
            data.to_excel(writer, sheet_name='Missing Integrator Transactions', startrow=count)
            count += 2
    wb.close()


def missing_OVA_transactions():
    for i in range(Ifirst_row+1, Ilast_row+1):
        for j in range(Ofirst_row+1, Olast_row+1):
            if Isheet['H' + str(i)].value == Osheet['B'+str(j)].value:
                matchFound = True
                break
            else:
                matchFound = False
        if matchFound == False:
            missing_OVA_list.append(Isheet['H' + str(i)].value)
    print(f"Missing OVA transactions: {missing_OVA_list}")
    wb = load_workbook('test_file.xlsx')
    ws3 = wb.create_sheet('Missing OVA Transactions')
    ws3.title = 'Missing OVA Transactions'
    file = pd.read_excel('MTN KR Debit INT_01_Oct_22.xlsx')
    count = 0
    for transaction in missing_OVA_list:
        data = file[file['IntegratorTransId'] == transaction]
        with pd.ExcelWriter('test_file.xlsx', mode='a', engine="openpyxl", if_sheet_exists='overlay') as writer:
            data.to_excel(writer, sheet_name='Missing OVA Transactions', startrow=count)
            count += 2
    wb.close()


Owb.close()
Iwb.close()


missing_OVA_transactions()
missing_INT_transactions()
