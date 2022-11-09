from openpyxl import load_workbook
import pandas as pd

wb = load_workbook('test_file.xlsx')
ws1 = wb.create_sheet("Duplicates")
ws1.title = 'Duplicates'
sheet = wb['MTN KR Debit Metabase_01_Oct_22']
file = pd.read_excel('test_file.xlsx')

first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row



list = []
checker_list = []
duplicate = []
count = 1
counter = 0
duplicates_value_sum = 0.00
for row in range(first_row+1, last_row+1):
    list.append(sheet['H'+str(row)].value)

for i in list:
    if i not in checker_list:
        checker_list.append(i)
    else:
        data = file[file['IntegratorTransId'] == i]
        duplicate.append(i)
        with pd.ExcelWriter('test_file.xlsx', mode='a', engine="openpyxl", if_sheet_exists='overlay') as writer:
            data.to_excel(writer, sheet_name='Duplicates', startrow=count)
            count += 3
wb.close()

wb = load_workbook('test_file.xlsx')
wb.active = wb['Duplicates']
dup_sheet = wb['Duplicates']

print(dup_sheet['B5'].value)
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row


for number in range(first_row+2, last_row+1,3):
    duplicates_value_sum = duplicates_value_sum + (dup_sheet['B' + str(number)].value)
    count =+ 3
    counter +=1
    print(duplicates_value_sum)


print(f"Number of duplicates: {counter}")
(dup_sheet['B' +str(last_row  +4)].value) = str(round(duplicates_value_sum, 2))
(dup_sheet['B' + str(last_row  +5)].value) = str(counter)

print(f"Duplicates: {duplicate}")


wb.close()
wb.save('test_file.xlsx')
