from openpyxl import load_workbook
from dateTime import yesterday
from matchCheck import *
# SLYDEPAY01 OVA
#
#
wb = load_workbook('SLYDEPAY01_OVA_01_Oct_22.xlsx')
sheet = wb['SLYDEPAY01_01_Oct_22']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

MIGS_01_OVA_Volume = (last_row + 1) - (first_row + 4)
OVA_VOLUME1 = MIGS_01_OVA_Volume
print('MIGS_01_OVA_Volume: ' + str(MIGS_01_OVA_Volume))

MIGS_01_OVA_Sum = 0.00

for i in range(first_row + 4, last_row+1):
    MIGS_01_OVA_Sum = MIGS_01_OVA_Sum + sheet['N' + str(i)].value

print('MIGS_01_OVA_Sum: ' + str(MIGS_01_OVA_Sum))
OVA_VALUE1 = MIGS_01_OVA_Sum
wb.close()
#
#
#
#  SLYDEPAY01 INT

wb = load_workbook('MIGS_INT_01_Oct_22.xlsx')
sheet = wb['MIGS Metabase_01_Oct_22']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

MIGS_01_INT_Volume = (last_row + 1) - (first_row + 1)
INT_VOLUME1 = MIGS_01_INT_Volume
print('MIGS_01_INT_Volume: ' + str(MIGS_01_INT_Volume))

MIGS_01_INT_Sum = 0.00

for i in range(first_row+1, last_row+1):
    MIGS_01_INT_Sum = MIGS_01_INT_Sum + sheet['H' + str(i)].value

print('MIGS_01_INT_Sum: ' + str(MIGS_01_INT_Sum))
INT_VALUE1 = MIGS_01_INT_Sum
wb.close()

#
#
#
#  SLYDEPULL PROMPTS OVA

wb = load_workbook('MTN Prompt_OVA_01_Oct_22.xlsx')
sheet = wb['MTN Prompt_01_Oct_22']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

MTN_PROMPT_OVA_Volume = (last_row + 1) - (first_row + 1)
OVA_VOLUME2 = MTN_PROMPT_OVA_Volume
print('MTN_PROMPT_OVA_Volume: ' + str(MTN_PROMPT_OVA_Volume))

MTN_PROMPT_OVA_Sum = 0.00

for i in range(first_row + 1, last_row+1):
    MTN_PROMPT_OVA_Sum = MTN_PROMPT_OVA_Sum + sheet['O' + str(i)].value

print('MTN_PROMPT_OVA_Sum: ' + str(MTN_PROMPT_OVA_Sum))
OVA_VALUE2 = MTN_PROMPT_OVA_Sum
wb.close()
#
#
#
# SLYDEPULL PROMPT INT

wb = load_workbook('MTN Prompt_INT_01_Oct_22.xlsx')
sheet = wb['MTN Prompt_01_Oct_22']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

MTN_PROMPT_INT_Volume = (last_row + 1) - (first_row + 1)
INT_VOLUME2 = MTN_PROMPT_INT_Volume
print('MTN_PROMPT_INT_Volume: ' + str(MTN_PROMPT_INT_Volume))

MTN_PROMPT_INT_Sum = 0.00

for i in range(first_row+1, last_row+1):
    MTN_PROMPT_INT_Sum = MTN_PROMPT_INT_Sum + sheet['O' + str(i)].value

print('MTN_PROMPT_INT_Sum: ' + str(MTN_PROMPT_INT_Sum))
INT_VALUE2 = MTN_PROMPT_INT_Sum

wb.close()
#
OVA_VOLUME3 = 0
OVA_VALUE3 = 0
INT_VOLUME3 = 0
INT_VALUE3 = 0
#
#
#
#  MTN PORTAL OVA

wb = load_workbook('MTN Cashout_OVA_01_Oct_22.xlsx')
sheet = wb['MTN Cashout_01_Oct_22']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

MTN_APPROVALS_OVA_Volume = (last_row + 1) - (first_row + 1)
OVA_VOLUME4 = MTN_APPROVALS_OVA_Volume
print('MTN_PORTAL_OVA_Volume: ' + str(MTN_APPROVALS_OVA_Volume))

MTN_APPROVALS_OVA_Sum = 0.00

for i in range(first_row+1, last_row+1):
    MTN_APPROVALS_OVA_Sum = MTN_APPROVALS_OVA_Sum + (-1 * sheet['O' + str(i)].value)

print('MTN_PORTAL_OVA_Sum: ' + str(MTN_APPROVALS_OVA_Sum))
OVA_VALUE4 = MTN_APPROVALS_OVA_Sum
wb.close()
#
#
#
INT_VOLUME4 = 0
INT_VALUE4 = 0
#
OVA_VOLUME5 = 0
OVA_VALUE5 = 0
INT_VALUE5 = 0
INT_VOLUME5 = 0
#
#
#
OVA_VOLUME6 = 0
OVA_VALUE6 = 0
INT_VALUE6 = 0
INT_VOLUME6 = 0
#
#
#  SP VODAFONE CASHIN OVA

wb = load_workbook('Vodafone Cashin_OVA_01_Oct_22 - Account Statement.xlsx')
sheet = wb['Vodafone Cashin_01_Oct_22.xls -']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

VODA_CASHIN_OVA_Volume = (last_row + 1) - (first_row + 6)
OVA_VOLUME7 = VODA_CASHIN_OVA_Volume
print('VODA_CASHOUT_OVA_Volume: ' + str(VODA_CASHIN_OVA_Volume))

VODA_Cashin_OVA_Sum = 0.00

for i in range(first_row+6, last_row+1):
    VODA_Cashin_OVA_Sum = VODA_Cashin_OVA_Sum + sheet['G' + str(i)].value

print('VODA_Cashout_OVA_Sum: ' + str(VODA_Cashin_OVA_Sum))
OVA_VALUE7 = VODA_Cashin_OVA_Sum
wb.close()
#
#
INT_VOLUME7 = 0
INT_VALUE7 = 0
#
#  VODA SP CASHOUT OVA

wb = load_workbook('Vodafone Cashout_OVA_01_Oct_22- Account Statement.xlsx')
sheet = wb['Vodafone Cashout_01_Oct_22.xls ']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

VODA_CASHOUT_OVA_Volume = (last_row + 1) - (first_row + 6)
OVA_VOLUME8 = VODA_CASHOUT_OVA_Volume
print('VODA_CASHOUT_OVA_Volume: ' + str(VODA_CASHOUT_OVA_Volume))

VODA_Cashout_OVA_Sum = 0.00

for i in range(first_row+6, last_row+1):
    VODA_Cashout_OVA_Sum = VODA_Cashout_OVA_Sum + (-1 * sheet['H' + str(i)].value)

print('VODA_Cashout_OVA_Sum: ' + str(VODA_Cashout_OVA_Sum))
OVA_VALUE8 = VODA_Cashout_OVA_Sum
wb.close()
#
INT_VOLUME8 = 0
INT_VALUE8 = 0
#
#
#  STANBIC FI CR OVA

wb = load_workbook('Stanbic FI CreditDebit_OVA_01_Oct_22.xlsx')
sheet = wb['Sheet1']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

Stanbic_FI_Credit_OVA_Volume = 0

Stanbic_FI_Credit_OVA_Sum = 0.00

for i in range(first_row+1, last_row+1):
    if sheet['D' + str(i)].value == 'C':
        Stanbic_FI_Credit_OVA_Volume = Stanbic_FI_Credit_OVA_Volume + 1
        Stanbic_FI_Credit_OVA_Sum = Stanbic_FI_Credit_OVA_Sum + sheet['F' + str(i)].value

print('Stanbic_FI_CREDIT_OVA_Volume: ' + str(Stanbic_FI_Credit_OVA_Volume))
print('Stanbic_FI_CREDIT_OVA_Sum: ' + str(Stanbic_FI_Credit_OVA_Sum))
OVA_VOLUME9 = Stanbic_FI_Credit_OVA_Volume
OVA_VALUE9 = Stanbic_FI_Credit_OVA_Sum
wb.close()
#
#
#
#  STANBIC FI CREDIT INT

wb = load_workbook('Stanbic FI Credit INT_01_Oct_22.xlsx')
sheet = wb['Stanbic FI Credit INT_01_Oct_22']

first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

Stanbic_FI_Credit_INT_Volume = (last_row + 1) - (first_row + 1)
Stanbic_FI_Credit_INT_Sum = 0.00
INT_VOLUME9 = Stanbic_FI_Credit_INT_Volume

for i in range(first_row+1, last_row+1):
    Stanbic_FI_Credit_INT_Sum = Stanbic_FI_Credit_INT_Sum + sheet['H' + str(i)].value

print('Stanbic_FI_Credit_INT_Volume: ' + str(Stanbic_FI_Credit_INT_Volume))
print('Stanbic_FI_Credit_INT_Sum: ' + str(Stanbic_FI_Credit_INT_Sum))
INT_VALUE9 = Stanbic_FI_Credit_INT_Sum
wb.close()
#
#
OVA_VOLUME10 = 0
OVA_VALUE10 = 0
INT_VALUE10 = 0
INT_VOLUME10 = 0
#
#
OVA_VALUE11 = 0
OVA_VOLUME11 = 0
#  GIP INT
wb = load_workbook('GIP INT_01_Oct_22.xlsx')
sheet = wb['GIP Metabase_01_Oct_22']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

GIP_INT_Volume = (last_row + 1) - (first_row + 1)
INT_VOLUME11 = GIP_INT_Volume
print('GIP_INT_Volume: ' + str(GIP_INT_Volume))

GIP_INT_Sum = 0.00

for i in range(first_row+1, last_row+1):
    GIP_INT_Sum = GIP_INT_Sum + sheet['I' + str(i)].value

print('GIP_INT_Sum: ' + str(GIP_INT_Sum))
INT_VALUE11 = GIP_INT_Sum
wb.close()
#
#
#
# BB MIG_OVA

wb8 = load_workbook('SLYDEPAY08_01_Sep_22.xlsx')
wb9 = load_workbook('SLYDEPAY09_01_Oct_22.xlsx')

sheet8 = wb8['SLYDEPAY08']
sheet9 = wb9['SLYDEPAY09']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col8 = wb8.active.min_column
last_col8 = wb8.active.max_column
first_row8 = wb8.active.min_row
last_row8 = wb8.active.max_row

first_col9 = wb9.active.min_column
last_col9 = wb9.active.max_column
first_row9 = wb9.active.min_row
last_row9 = wb9.active.max_row

BB_MIG8_Volume = (last_row8 + 1) - (first_row8 + 4)

BB_MIG8_Sum = 0.00

for i in range(first_row8 + 4, last_row8+1):
    BB_MIG8_Sum = BB_MIG8_Sum + sheet8['N' + str(i)].value

BB_MIG9_Volume = (last_row9 + 1) - (first_row9 + 4)
BB_MIG9_Sum = 0.00

for i in range(first_row9 + 4, last_row9 + 1):
    BB_MIG9_Sum = BB_MIG9_Sum + sheet9['K' + str(i)].value

BB_MIG_VOLUME = BB_MIG9_Volume + BB_MIG8_Volume
OVA_VOLUME12 = BB_MIG_VOLUME
BB_MIG_SUM = BB_MIG9_Sum + BB_MIG8_Sum
OVA_VALUE12 = BB_MIG_SUM

print('BB_MIG_OVA_VOLUME: ' + str(BB_MIG_VOLUME))
print('BB_MIG_OVA_Sum: ' + str(BB_MIG_SUM))
wb8.close()
wb9.close()
#
#
#
# BB MIG_INT

wb = load_workbook('MiGS_8_9_INT.xlsx')
sheet = wb['MiGS8_9_INT']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

MIGS_INT_Sum = 0.00
MIGS_INT_VOLUME = 0

for i in range(first_row+1, last_row+1):
    if sheet['P' + str(i)].value == 'CONFIRMED':
        MIGS_INT_VOLUME = MIGS_INT_VOLUME + 1
        MIGS_INT_Sum = MIGS_INT_Sum + sheet['E' + str(i)].value

print('MIGS_INT_VOLUME: ' + str(MIGS_INT_VOLUME))
INT_VOLUME12 = MIGS_INT_VOLUME
print('MIGS_INT_Sum: ' + str(MIGS_INT_Sum))
INT_VALUE12 = MIGS_INT_Sum
wb.close()
#
#
#
#  MPGS OVA

wb = load_workbook('MGPS_OVA.xlsx')
sheet = wb['MPGS_01_Oct_22']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

MGPS_OVA_Volume = (last_row + 1) - (first_row + 1)
OVA_VOLUME13 = MGPS_OVA_Volume
print('MGPS_OVA_Volume: ' + str(MGPS_OVA_Volume))

MGPS_OVA_Sum = 0.00

for i in range(first_row+1, last_row+1):
    MGPS_OVA_Sum = MGPS_OVA_Sum + sheet['I' + str(i)].value

print('MGPS_OVA_Sum: ' + str(MGPS_OVA_Sum))
OVA_VALUE13 = MGPS_OVA_Sum
wb.close()
#
#
INT_VOLUME13 = 0
INT_VALUE13 = 0
#
# MTN KR CREDIT OVA
wb = load_workbook('MTN KR Credit OVA_01_Oct_22.xlsx')
sheet = wb['MTN KR Credit Metabase_01_Oct_2']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

MTN_KR_Credit_OVA_Volume = (last_row + 1) - (first_row + 1)
OVA_VOLUME14 = MTN_KR_Credit_OVA_Volume
print('MTN_KR_CREDIT_OVA_Volume: ' + str(MTN_KR_Credit_OVA_Volume))

MTN_KR_Credit_OVA_Sum = 0.00

for i in range(first_row+1, last_row+1):
    MTN_KR_Credit_OVA_Sum = MTN_KR_Credit_OVA_Sum + (-1 * sheet['O' + str(i)].value)

print('MTN_KR_CREDIT_OVA_Sum: ' + str(MTN_KR_Credit_OVA_Sum))
OVA_VALUE14 = MTN_KR_Credit_OVA_Sum
wb.close()
#
#
#
# MTN KR CREDIT INT

wb = load_workbook('MTN KR Credit_INT_01_Oct_22.xlsx')
sheet = wb['MTN KR Credit_01_Oct_22']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

MTN_KR_Credit_INT_Volume = (last_row + 1) - (first_row + 1)
INT_VOLUME14 = MTN_KR_Credit_INT_Volume
print('MTN_KR_Credit_INT_Volume: ' + str(MTN_KR_Credit_INT_Volume))

MTN_KR_Credit_INT_Sum = 0.00

for i in range(first_row+1, last_row+1):
    MTN_KR_Credit_INT_Sum = MTN_KR_Credit_INT_Sum + sheet['B' + str(i)].value

print('MTN_KR_Credit_INT_Sum: ' + str(MTN_KR_Credit_INT_Sum))
INT_VALUE14 = MTN_KR_Credit_INT_Sum
wb.close()
#
#
#
# MTN KR DEBIT OVA

wb = load_workbook('MTN KR Debit_OVA_01_Oct_22.xlsx')
sheet = wb['Sheet1']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

MTN_KR_Debit_OVA_Volume = (last_row + 1) - (first_row + 1)
OVA_VOLUME15 = MTN_KR_Debit_OVA_Volume
print('MTN_KR_Debit_OVA_Volume: ' + str(MTN_KR_Debit_OVA_Volume))

MTN_KR_Debit_OVA_Sum = 0.00

for i in range(first_row+1, last_row+1):
    MTN_KR_Debit_OVA_Sum = MTN_KR_Debit_OVA_Sum + sheet['O' + str(i)].value

print('MTN_KR_Debit_OVA_Sum: ' + str(MTN_KR_Debit_OVA_Sum))
OVA_VALUE15 = MTN_KR_Debit_OVA_Sum
wb.close()
#
#
#
# MTN KR DEBIT INT

wb = load_workbook('MTN KR Debit INT_01_Oct_22.xlsx')
sheet = wb['MTN KR Debit Metabase_01_Oct_22']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

MTN_KR_Debit_INT_Volume = (last_row + 1) - (first_row + 1)
INT_VOLUME15 = MTN_KR_Debit_INT_Volume
print('MTN_KR_Debit_INT_Volume: ' + str(MTN_KR_Debit_INT_Volume))

MTN_KR_Debit_INT_Sum = 0.00

for i in range(first_row+1, last_row):
    print(sheet['A' + str(i)].value)
    MTN_KR_Debit_INT_Sum = MTN_KR_Debit_INT_Sum + sheet['A' + str(i)].value

print('MTN_KR_Debit_INT_Sum: ' + str(MTN_KR_Debit_INT_Sum))
INT_VALUE15 = MTN_KR_Debit_INT_Sum
wb.close()
#
#
OVA_VALUE16 = 0
OVA_VOLUME16 = 0
INT_VOLUME16 = 0
INT_VALUE16 = 0
#
#
#
OVA_VALUE17 = 0
OVA_VOLUME17 = 0
INT_VOLUME17 = 0
INT_VALUE17 = 0
# VODA KR CASHIN OVA

wb = load_workbook('Vodafone KR Cashin_OVA_01_Oct_22.xlsx')
sheet = wb['Account Statement']

# DEFINE MAX AND MIN COLUMNS AND ROWS
first_col = wb.active.min_column
last_col = wb.active.max_column
first_row = wb.active.min_row
last_row = wb.active.max_row

VODA_CASHIN_KR_OVA_Volume = (last_row + 1) - (first_row + 6)
OVA_VOLUME18 = VODA_CASHIN_KR_OVA_Volume
print('VODA_CASHOUT_OVA_Volume: ' + str(VODA_CASHIN_KR_OVA_Volume))

VODA_KR_Cashin_OVA_Sum = 0.00

for i in range(first_row+6, last_row+1):
    VODA_KR_Cashin_OVA_Sum = VODA_KR_Cashin_OVA_Sum + sheet['G' + str(i)].value

print('VODA_Cashout_OVA_Sum: ' + str(VODA_KR_Cashin_OVA_Sum))
OVA_VALUE18 = VODA_KR_Cashin_OVA_Sum
wb.close()
#
#
INT_VOLUME18 = 0
INT_VALUE18 = 0
OVA_VOLUME19 = 0
OVA_VALUE19 = 0
INT_VALUE19 = 0
INT_VOLUME19 = 0
#
#
#

# VARIANCE COUNTS & VALUES
VARIANCE_VOLUME1 = abs(INT_VOLUME1 - OVA_VOLUME1)
VARIANCE_VOLUME2 = abs(INT_VOLUME2 - OVA_VOLUME2)
VARIANCE_VOLUME3 = abs(INT_VOLUME3 - OVA_VOLUME3)
VARIANCE_VOLUME4 = abs(INT_VOLUME4 - OVA_VOLUME4)
VARIANCE_VOLUME5 = abs(INT_VOLUME5 - OVA_VOLUME5)
VARIANCE_VOLUME6 = abs(INT_VOLUME6 - OVA_VOLUME6)
VARIANCE_VOLUME7 = abs(INT_VOLUME7 - OVA_VOLUME7)
VARIANCE_VOLUME8 = abs(INT_VOLUME8 - OVA_VOLUME8)
VARIANCE_VOLUME9 = abs(INT_VOLUME9 - OVA_VOLUME9)
VARIANCE_VOLUME10 = abs(INT_VOLUME10 - OVA_VOLUME10)
VARIANCE_VOLUME11 = abs(INT_VOLUME11 - OVA_VOLUME11)
VARIANCE_VOLUME12 = abs(INT_VOLUME12 - OVA_VOLUME12)
VARIANCE_VOLUME13 = abs(INT_VOLUME13 - OVA_VOLUME13)
VARIANCE_VOLUME14 = abs(INT_VOLUME14 - OVA_VOLUME14)
VARIANCE_VOLUME15 = abs(INT_VOLUME15 - OVA_VOLUME15)
VARIANCE_VOLUME16 = abs(INT_VOLUME16 - OVA_VOLUME16)
VARIANCE_VOLUME17 = abs(INT_VOLUME17 - OVA_VOLUME17)
VARIANCE_VOLUME18 = abs(INT_VOLUME18 - OVA_VOLUME18)
VARIANCE_VOLUME19 = abs(INT_VOLUME19 - OVA_VOLUME19)

VARIANCE_VOLUME_LIST1 = [VARIANCE_VOLUME1, VARIANCE_VOLUME2, VARIANCE_VOLUME3, VARIANCE_VOLUME4, VARIANCE_VOLUME5, VARIANCE_VOLUME6, VARIANCE_VOLUME7, VARIANCE_VOLUME8, VARIANCE_VOLUME9, VARIANCE_VOLUME10, VARIANCE_VOLUME11]
VARIANCE_VOLUME_LIST2 = [VARIANCE_VOLUME12, VARIANCE_VOLUME13, VARIANCE_VOLUME14, VARIANCE_VOLUME15, VARIANCE_VOLUME16, VARIANCE_VOLUME17, VARIANCE_VOLUME18, VARIANCE_VOLUME19]

VARIANCE_VALUE1 = abs(INT_VALUE1 - OVA_VALUE1)
VARIANCE_VALUE2 = abs(INT_VALUE2 - OVA_VALUE2)
VARIANCE_VALUE3 = abs(INT_VALUE3 - OVA_VALUE3)
VARIANCE_VALUE4 = abs(INT_VALUE4 - OVA_VALUE4)
VARIANCE_VALUE5 = abs(INT_VALUE5 - OVA_VALUE5)
VARIANCE_VALUE6 = abs(INT_VALUE6 - OVA_VALUE6)
VARIANCE_VALUE7 = abs(INT_VALUE7 - OVA_VALUE7)
VARIANCE_VALUE8 = abs(INT_VALUE8 - OVA_VALUE8)
VARIANCE_VALUE9 = abs(INT_VALUE9 - OVA_VALUE9)
VARIANCE_VALUE10 = abs(INT_VALUE10 - OVA_VALUE10)
VARIANCE_VALUE11 = abs(INT_VALUE11 - OVA_VALUE11)
VARIANCE_VALUE12 = abs(INT_VALUE12 - OVA_VALUE12)
VARIANCE_VALUE13 = abs(INT_VALUE13 - OVA_VALUE13)
VARIANCE_VALUE14 = abs(INT_VALUE14 - OVA_VALUE14)
VARIANCE_VALUE15 = abs(INT_VALUE15 - OVA_VALUE15)
VARIANCE_VALUE16 = abs(INT_VALUE16 - OVA_VALUE16)
VARIANCE_VALUE17 = abs(INT_VALUE17 - OVA_VALUE17)
VARIANCE_VALUE18 = abs(INT_VALUE18 - OVA_VALUE18)
VARIANCE_VALUE19 = abs(INT_VALUE19 - OVA_VALUE19)

VARIANCE_VALUE_LIST1 = [VARIANCE_VALUE1, VARIANCE_VALUE2, VARIANCE_VALUE3, VARIANCE_VALUE4, VARIANCE_VALUE5, VARIANCE_VALUE6, VARIANCE_VALUE7, VARIANCE_VALUE8, VARIANCE_VALUE9, VARIANCE_VALUE10, VARIANCE_VALUE11]
VARIANCE_VALUE_LIST2 = [VARIANCE_VALUE12, VARIANCE_VALUE13, VARIANCE_VALUE14, VARIANCE_VALUE15, VARIANCE_VALUE16, VARIANCE_VALUE17, VARIANCE_VALUE18, VARIANCE_VALUE19]
# update sheet
fwb = load_workbook('Reconciliations 2022.xlsx')
fsheet = fwb['OCT (2)']

OVA_VOLUME_LIST1 = [OVA_VOLUME1, OVA_VOLUME2, OVA_VOLUME3, OVA_VOLUME4, OVA_VOLUME5, OVA_VOLUME6, OVA_VOLUME7, OVA_VOLUME8, OVA_VOLUME9, OVA_VOLUME10, OVA_VOLUME11]
i = 2
for volume in OVA_VOLUME_LIST1:
    fsheet['F' + str(i)] = volume
    i = i + 1

OVA_VOLUME_LIST2 = [OVA_VOLUME12, OVA_VOLUME13, OVA_VOLUME14, OVA_VOLUME15, OVA_VOLUME16, OVA_VOLUME17, OVA_VOLUME18, OVA_VOLUME19]
i = 23
for volume in OVA_VOLUME_LIST2:
    fsheet['F' + str(i)] = volume
    i = i + 1

OVA_VALUE_LIST1 = [OVA_VALUE1, OVA_VALUE2, OVA_VALUE3, OVA_VALUE4, OVA_VALUE5, OVA_VALUE6, OVA_VALUE7, OVA_VALUE8, OVA_VALUE9, OVA_VALUE10, OVA_VALUE11]
OVA_VALUE_LIST2 = [OVA_VALUE12, OVA_VALUE13, OVA_VALUE14, OVA_VALUE15, OVA_VALUE16, OVA_VALUE17, OVA_VALUE18, OVA_VALUE19]

i = 2
for volume in OVA_VALUE_LIST1:
    fsheet['G' + str(i)] = volume
    i = i + 1

i = 23
for volume in OVA_VALUE_LIST2:
    fsheet['G' + str(i)] = volume
    i = i + 1

INT_VOLUME_LIST1 = [INT_VOLUME1, INT_VOLUME2, INT_VOLUME3, INT_VOLUME4, INT_VOLUME5, INT_VOLUME6, INT_VOLUME7, INT_VOLUME8, INT_VOLUME9, INT_VOLUME10, INT_VOLUME11]
INT_VOLUME_LIST2 = [INT_VOLUME12, INT_VOLUME13, INT_VOLUME14, INT_VOLUME15, INT_VOLUME16, INT_VOLUME17, INT_VOLUME18, INT_VOLUME19]


i = 2
for volume in INT_VOLUME_LIST1:
    fsheet['H'+str(i)] = volume
    i += 1

i = 23
for volume in INT_VOLUME_LIST2:
    fsheet['H'+str(i)] = volume
    i += 1

INT_VALUE_LIST1 = [INT_VALUE1, INT_VALUE2, INT_VALUE3, INT_VALUE4, INT_VALUE5, INT_VALUE6, INT_VALUE7, INT_VALUE8, INT_VALUE9, INT_VALUE10, INT_VALUE11]
INT_VALUE_LIST2 = [INT_VALUE12, INT_VALUE13, INT_VALUE14, INT_VALUE15, INT_VALUE16, INT_VALUE17, INT_VALUE18, INT_VALUE19]

i = 2
for value in INT_VALUE_LIST1:
    fsheet['I'+str(i)] = value
    i += 1

i = 23
for value in INT_VALUE_LIST2:
    fsheet['I' + str(i)] = value
    i += 1

i = 2
for count in VARIANCE_VOLUME_LIST1:
    fsheet['J' + str(i)] = count
    i += 1

i = 23
for count in VARIANCE_VOLUME_LIST2:
    fsheet['J' + str(i)] = count
    i += 1

i = 2
for value in VARIANCE_VALUE_LIST1:
    fsheet['K' + str(i)] = value
    i += 1

i = 23
for value in VARIANCE_VALUE_LIST2:
    fsheet['K' + str(i)] = value
    i += 1
fwb.save('test.xlsx')
