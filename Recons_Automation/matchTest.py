from openpyxl import load_workbook

Owb= load_workbook('Book1.xlsx')
Osheet = Owb['Sheet1']

Iwb = load_workbook('Book2.xlsx')
Isheet = Iwb['Sheet1']

Ofirst_col = Owb.active.min_column
Olast_col = Owb.active.max_column
Ofirst_row = Owb.active.min_row
Olast_row = Owb.active.max_row

Ifirst_col = Iwb.active.min_column
Ilast_col = Iwb.active.max_column
Ifirst_row = Iwb.active.min_row
Ilast_row = Iwb.active.max_row

matchFound = False
missing_list= []
Mismatch = 0


for i in range(Ofirst_row, Olast_row+1):
    for j in range(Ifirst_row, Ilast_row+1):
        if Osheet['A'+ str(i)].value == Isheet['A'+str(j)].value:
            matchFound = True
            break
        else:
            matchFound=False
    if matchFound == False:
        missing_list.append(Osheet['A'+ str(i)].value)

print(f"In book1 but not in book 2 {missing_list}")