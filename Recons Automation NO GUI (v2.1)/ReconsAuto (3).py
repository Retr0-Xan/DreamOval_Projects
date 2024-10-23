# version 2.0.1
import pandas as pd
from typing import Literal, Tuple
from openpyxl import load_workbook

# from dateTime import yesterday, GIPdate, current_month, recons_yesterday
from datetime import datetime
from datetime import date
from datetime import timedelta
import os

tx_id_col_names = [
    "integratorTransId",
    "IntegratorTransId",
    "Transaction Id",
    "TransId",
    "External Transaction Id",
    "BillerTransId",
    "Id",
    "External Payment Request â†’ Institution Trans ID",
    "Merchant Transaction Reference",
    "REMARKS2",
    "External Payment Request → Institution Trans ID",
    "Order Code",
    "Order ID",
    "Integrator Trans ID",
    "REFERENCE_NUMBER",
]
amount_col_names = [
    "Amount",
    "amount",
    "Paid in",
    "Paid In",
    "Withdrawn",
    "AMOUNT",
    "Amount ($)",
    "Actual Amount ($)",
    "Transaction Amount (GHC.)",
    "TRANSACTION_AMOUNT",
    "Transaction Amount (amount only)",
    "Order Amount (amount only)",
    "Real Total",
    "Total",
]


def check_for_file(file_name):
    # This function is called when calling the run_recons fucntion.
    # It checks if a particular file can be found in the directory and returns the file name else returns None
    if file_name in os.listdir():
        return file_name
    else:
        return None


def get_date(
    date: date, format: Literal["gip"] or Literal["normal"] or Literal["recons"]
):
    """
    Formats a given date according to the specified format.

    Parameters:
    date (date): A datetime.date object representing the date to be formatted.
    format (Literal["gip"] or Literal["normal"] or Literal["recons"]): The format to use for formatting the date.
        - "gip": Format date as "YYYYMMDD" (e.g., 20231201).
        - "recons": Format date as "YYYY-MM-DD" (e.g., 2023-12-01).
        - "normal": Format date as day_month_year (e.g., 1_Jan_23).

    Returns:
    str: The formatted date string based on the specified format.

    Example:
    input_date = date(2023, 12, 1)

    # Format as "YYYYMMDD"
    gip_format = get_date(input_date, "gip")
    print(gip_format)
    # Output: "20231201"

    # Format as "YYYY-MM-DD"
    recons_format = get_date(input_date, "recons")
    print(recons_format)
    # Output: "2023-12-01"

    # Format as "normal"
    normal_format = get_date(input_date, "normal")
    print(normal_format)
    # Output: "1 Jan_23"

    Note:
    - This function allows formatting a date in three different formats based on the 'format' parameter.
    - 'gip' format represents the date as "YYYYMMDD".
    - 'recons' format represents the date as "YYYY-MM-DD".
    - 'normal' format represents the date as day_month_year (e.g., 1_Jan_23).
    """
    if format == "gip":
        return date.strftime("%Y%m%d")  # Output date in format: 20231201
    elif format == "recons":
        return date.strftime("%Y-%m-%d")  # Output data in recons: 2023-01-01
    else:
        return date.strftime("%d %b_%y")  # Output date in format: 1_Jan_23


def get_write_double_ova_val(
    ova_file: str or None,
    num_lines_of_header: Tuple[int, int],
    alt_recons_name: str,
    file_output_name: str,
):
    """
    # Note: This function is only used in BB MIG recons.

    The get_write_double_ova_val function processes data from two OVA files, extracts relevant information, and generates a consolidated report.
    It also outputs the processed data into a new Excel file for further analysis.

    Parameters
    ----------
        -ova_files (Tuple[str or None, str or None]): A tuple containing the file paths of the two OVA files. If a file path is None,the function processes the available file.
        -num_lines_of_header (Tuple[int, int]): A tuple containing the number of lines to skip as headers for each OVA file.
        -alt_recons_name (str): A string representing an alternative name for the reconciliation file.
        -file_output_name (str): A string representing the name of the output file.

    Process
    -------
        -The function reads the data from the two OVA files into separate DataFrames (ova_df1 and ova_df2).
        -It calculates the total volume and value of the OVA transactions by summing the relevant columns.
        -The function prints the OVA volume and value for analysis.
        -The processed data from the first OVA file is written into a new Excel file.
        -If one of the OVA files is missing (None), the function processes the available file and returns the name of the generated reconciliation file.
    """

    ova_df1 = pd.read_excel(ova_file, skiprows=num_lines_of_header[0])

    ova_volume = len(ova_df1)
    ova_volumes[21] = ova_volume
    recons_file = (
        f"{alt_recons_name} - Recons.xlsx"
        if ova_file is None
        else f"{ova_file[:-5]} - Recons.xlsx"
    )
    df1_amount_col = ""
    for name in amount_col_names:
        if name in ova_df1.columns:
            if not ova_df1[name].isna().all():
                df1_amount_col = name
                break

    ova_value = ova_df1[df1_amount_col].abs().sum()
    ova_values[21] = ova_value
    print(f"{file_output_name} OVA VOLUME : {ova_volume}")
    print(f"{file_output_name} OVA VALUE : {ova_value}")

    with pd.ExcelWriter(
        recons_file, engine="openpyxl", mode="w"
    ) as writer:  # specify new file name to write to
        ova_df1.to_excel(
            writer, sheet_name="Sheet1", index=False
        )  # save original data into first sheet of new file
        return recons_file


def get_write_double_int_val(
    int_files: Tuple[str or None, str or None],
    num_lines_of_header: Tuple[int, int],
    alt_recons_name: str,
    file_output_name: str,
    recons_file: str,
):
    # see get_write_double_ova_val function
    try:
        int_df1 = pd.read_excel(int_files[0], skiprows=num_lines_of_header[0])
        int_df2 = pd.read_excel(int_files[1], skiprows=num_lines_of_header[1])
        int_df2 = int_df2.loc[int_df2["Status"] == "CONFIRMED"]

        int_volume = len(int_df1) + len(int_df2)

        df1_amount_col = ""
        for name in amount_col_names:
            if name in int_df1.columns:
                if not int_df1[name].isna().all():
                    df1_amount_col = name
                    break

        df2_amount_col = ""
        for name in amount_col_names:
            if name in int_df2.columns:
                if not int_df2[name].isna().all():
                    df2_amount_col = name
                    break

        int_value = (
            int_df1[df1_amount_col].abs().sum() + int_df2[df2_amount_col].abs().sum()
        )
        print(f"{file_output_name} INT VOLUME : {int_volume}")
        print(f"{file_output_name} INT VALUE : {int_value}")
        dup, dup_val = find_duplicates(int_df1)
        write_duplicate_data(
            amount_col_name=df1_amount_col, df=dup, value=dup_val, file_name=recons_file
        )
        dup2, dup_val = find_duplicates(int_df2)
        write_duplicate_data(
            amount_col_name=df2_amount_col,
            df=dup2,
            value=dup_val,
            file_name=recons_file,
            last_row_index=len(dup) + 3,
        )
        return int_volume, int_value
    except ValueError:
        try:
            int_files = tuple(item for item in int_files if item is not None)
            if int_files[0] == f"MPGS KC{yesterday}.xlsx":
                int_df1 = pd.read_excel(int_files[0], skiprows=num_lines_of_header[0])
                int_volume = len(int_df1)

                amount_col = ""
                for name in amount_col_names:
                    if name in int_df1.columns:
                        if not int_df1[name].isna().all():
                            amount_col = name
                            break

                int_value = int_df1[amount_col].abs().sum()
                print(f"{file_output_name} INT VOLUME : {int_volume}")
                print(f"{file_output_name} INT VALUE : {int_value}")
                dup, dup_val = find_duplicates(int_df1)
                write_duplicate_data(
                    amount_col_name=amount_col,
                    df=dup,
                    value=dup_val,
                    file_name=recons_file,
                )
                return int_volume, int_value
            elif int_files[0] == f"MPGS_trn{yesterday}.xlsx":
                int_df1 = pd.read_excel(int_files[0], skiprows=num_lines_of_header[0])
                int_df1 = int_df1.loc[int_df1["Status"] == "CONFIRMED"]
                int_volume = len(int_df1)

                amount_col = ""
                for name in amount_col_names:
                    if name in int_df1.columns:
                        if not int_df1[name].isna().all():
                            amount_col = name
                            break

                int_value = int_df1[amount_col].abs().sum()
                print(f"{file_output_name} INT VOLUME : {int_volume}")
                print(f"{file_output_name} INT VALUE : {int_value}")
                dup, dup_val = find_duplicates(int_df1)
                write_duplicate_data(
                    amount_col_name=amount_col,
                    df=dup,
                    value=dup_val,
                    file_name=recons_file,
                )
                return int_volume, int_value
        except:
            print("INT FILE NOT FOUND")


def update_recons_sheet():
    fwb = load_workbook("Reconciliations 2024.xlsx")

    if datetime.today().day == 1:
        today = date.today()
        prev_month = (
            today.replace(month=today.month - 1)
            if today.month > 1
            else today.replace(month=12, year=today.year - 1)
        )
        prev_month_abbr = prev_month.strftime("%b").upper()
        fsheet = fwb[prev_month_abbr]
        fwb.active = fwb[prev_month_abbr]
    else:
        fsheet = fwb[current_month]
        fwb.active = fwb[current_month]

    first_col = fwb.active.min_column  # type: ignore
    last_col = fwb.active.max_column  # type: ignore
    first_row = fwb.active.min_row  # type: ignore
    last_row = fwb.active.max_row  # type: ignore

    global start_row
    for row in range(first_row + 1, last_row + 1):
        if str(fsheet["A" + str(row)].value) == recons_yesterday:
            start_row = row
            break
    print(f" start row = {start_row}")
    for row in range(0, 17):
        fsheet["E" + str(start_row)].value = ova_volumes[row]
        fsheet["F" + str(start_row)].value = abs(ova_values[row])
        fsheet["G" + str(start_row)].value = int_volumes[row]
        fsheet["H" + str(start_row)].value = abs(int_values[row])
        fsheet["M" + str(start_row)].value = dup_volumes[row]
        fsheet["N" + str(start_row)].value = abs(dup_values[row])
        start_row += 1
    fwb.close()
    fwb.save("Reconciliations 2024.xlsx")


def find_duplicates(int_df: pd.DataFrame):
    trans_id_col = ""
    for name in tx_id_col_names:
        if name in int_df.columns:
            trans_id_col = name
            break

    amount_col = ""
    for name in amount_col_names:
        if name in int_df.columns:
            if not int_df[name].isna().all():
                amount_col = name
                break

    if trans_id_col == "":
        raise Exception("No transaction id column found")

    unique_tx = {}
    duplicates_tx = []

    duplicate_value: float = 0

    for index, row in int_df.iterrows():
        tx_id = row[trans_id_col]
        if tx_id in unique_tx:  # duplicate
            duplicates_tx.append(row)
            duplicate_value += row[amount_col]
        else:
            unique_tx[tx_id] = index

    return pd.DataFrame(duplicates_tx), duplicate_value


def write_duplicate_data(
    amount_col_name: str,
    df: pd.DataFrame,
    value: float,
    file_name: str,
    last_row_index=None,
):
    if df.empty:
        return
    number_of_duplicates = len(df)
    amount_column = df.columns.get_loc(amount_col_name)
    if last_row_index is None:
        last_row_index = df.shape[0]
    empty_rows = pd.DataFrame(
        {col: [None] for col in df.columns},
        index=range(last_row_index, last_row_index + 3),
    )
    empty_rows.iat[-2, amount_column] = value
    empty_rows.iat[-1, amount_column] = number_of_duplicates

    df = pd.concat([df, empty_rows], ignore_index=True)

    mode = "a" if os.path.exists(file_name) else "w"
    with pd.ExcelWriter(file_name, engine="openpyxl", mode=mode) as writer:
        sheet_name = "Duplicates"
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def write_missing_ova_data(
    amount_col_name: str,
    df: pd.DataFrame,
    value: float,
    file_name: str,
    last_row_index=None,
):
    if df.empty:
        return
    # If last_row_index is not specified, start from the end of the existing data
    if last_row_index is None:
        last_row_index = df.shape[0]
    number_of_tx = len(df)
    empty_rows = pd.DataFrame(
        {col: [None] for col in df.columns},
        index=range(last_row_index, last_row_index + 3),
    )
    amount_column = df.columns.get_loc(amount_col_name)
    # Concatenate the empty rows with the original DataFrame
    df = pd.concat([df, empty_rows], ignore_index=True)
    # Update the number of tx and value in the new rows
    df.iat[last_row_index + 1, amount_column] = value
    df.iat[last_row_index + 2, amount_column] = number_of_tx
    with pd.ExcelWriter(file_name, engine="openpyxl", mode="a") as writer:
        sheet_name = "Missing OVA Transactions"
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def write_missing_int_data(
    amount_col_name: str,
    df: pd.DataFrame,
    value: float,
    file_name: str,
    last_row_index=None,
):
    if df.empty:
        return
    # If last_row_index is not specified, start from the end of the existing data
    if last_row_index is None:
        last_row_index = df.shape[0]
    number_of_tx = len(df)
    empty_rows = pd.DataFrame(
        {col: [None] for col in df.columns},
        index=range(last_row_index, last_row_index + 3),
    )
    amount_column = df.columns.get_loc(amount_col_name)
    # Concatenate the empty rows with the original DataFrame
    df = pd.concat([df, empty_rows], ignore_index=True)
    # Update the number of tx and value in the new rows
    df.iat[last_row_index + 1, amount_column] = value
    df.iat[last_row_index + 2, amount_column] = number_of_tx
    with pd.ExcelWriter(file_name, engine="openpyxl", mode="a") as writer:
        sheet_name = "Missing INT Transactions"
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def run_recons(
    file_names: Tuple[str or None, str or None],
    num_lines_of_header: Tuple[int, int],
    alt_recons_name: str,
    file_output_name: str,
    ova_id: str,
    int_id: str,
    alt_ova_id: str,
    alt_int_id: str or None = None,
    *,
    mb_service_name: str or None = None,
    mb_creditDebit_flag: str or None = None,
    mb_status_flag: str or None = None,
    ova_status_flag: str or None = None,
    ova_status_col: str or None = None,
    list_index: int,
):
    if file_names[0] is None and file_names[1] is None:
        return
    service_name_header = ""
    creditDebit_header = ""
    service_name_headers = ["ServiceName", "serviceName", "Service Name"]
    creditDebit_headers = [
        "creditDebitFlag",
        "CreditDebitFlag",
        "DEBITCREDIT",
        "Credit Debit Flag",
    ]
    # -------------------- OVA -------------------
    ova_file_name = file_names[0]  # name of the ova file
    int_file_name = file_names[1]  # name of the integrator file
    ova_header_lines = num_lines_of_header[0]
    int_header_lines = num_lines_of_header[1]
    recons_file = (
        f"{alt_recons_name} - Recons.xlsx"
        if ova_file_name is None
        else f"{ova_file_name[:-5]} - Recons.xlsx"
    )
    ova_file_df: pd.DataFrame or None = None
    int_file_df: pd.DataFrame or None = None

    if ova_file_name is not None:
        # put the data into a dataframe
        ova_file_df = pd.read_excel(ova_file_name, skiprows=ova_header_lines)
        ova_id_name = ova_id
        if ova_status_flag is not None:
            ova_file_df = ova_file_df.loc[
                ova_file_df[ova_status_col] == ova_status_flag
            ]

        for name in creditDebit_headers:
            if name in ova_file_df.columns:
                ova_file_df = ova_file_df[ova_file_df[name] == "C"]
                break
        ova_volume = len(ova_file_df)
        ova_volumes[list_index] = ova_volume
        print(
            f"{file_output_name}_OVA_Volume: {ova_volume}"
        )  # file_output_name is the name that shows for each channel as the script runs
        amount_col = ""
        for name in amount_col_names:
            if name in ova_file_df.columns:
                if not ova_file_df[name].isna().all():
                    amount_col = name
                    break  # check which of the formats the amount column is written in

        ova_value = ova_file_df[amount_col].abs().sum()
        ova_values[list_index] = ova_value
        print(f"{file_output_name} OVA_VALUE : {ova_value}")
        with pd.ExcelWriter(
            recons_file, engine="openpyxl", mode="w"
        ) as writer:  # specify new file name to write to
            ova_file_df.to_excel(
                writer, sheet_name="Sheet1", index=False
            )  # save original data into first sheet of new file

    # ----------------------- INTEGRATOR/ DUPLICATES --------------------
    if int_file_name is not None:
        int_file_df = pd.read_excel(int_file_name, skiprows=int_header_lines)
        int_id_name = int_id
        for name in service_name_headers:
            if name in int_file_df.columns:
                if not int_file_df[name].isna().all():
                    service_name_header = name
                    break
        for name in creditDebit_headers:
            if name in int_file_df.columns:
                if not int_file_df[name].isna().all():
                    creditDebit_header = name
                    break
        if (
            mb_service_name is not None
            and mb_creditDebit_flag is not None
            and mb_status_flag is None
        ):
            int_file_df = int_file_df.loc[
                (int_file_df[service_name_header] == mb_service_name)
                & (int_file_df[creditDebit_header] == mb_creditDebit_flag)
            ]
        elif (
            mb_service_name is not None
            and mb_creditDebit_flag is None
            and mb_status_flag is None
        ):
            int_file_df = int_file_df.loc[
                int_file_df[service_name_header] == mb_service_name
            ]
        if (
            mb_status_flag is not None
            and mb_service_name is None
            and mb_creditDebit_flag is None
        ):
            int_file_df = int_file_df.loc[int_file_df["Status"] == mb_status_flag]
        if int_file_df.empty:
            print(f"No transactions found for {file_output_name}. Confirm")
            return
        int_volume = len(int_file_df)
        int_volumes[list_index] = int_volume
        print(f"{file_output_name}_INT_Volume: {str(int_volume)}")
        amount_col = ""
        for name in amount_col_names:
            if name in int_file_df.columns:
                if not int_file_df[name].isna().all():
                    amount_col = name
                    break
        int_value = int_file_df[amount_col].abs().sum()
        int_values[list_index] = int_value

        print(f"{file_output_name} INT_VALUE : {int_value}")
        dup, dup_val = find_duplicates(int_file_df)
        print(f"Number of duplicates: {len(dup)}")
        print(f"Duplicates value: {dup_val}")
        dup_volumes[list_index] = len(dup)
        dup_values[list_index] = dup_val

        write_duplicate_data(
            amount_col_name=amount_col, df=dup, value=dup_val, file_name=recons_file
        )
    if ova_file_name == f"MPGS{yesterday}.xlsx":
        return
    # ---------------------- MISSING TRANSACTIONS -------------------------
    if ova_file_df is not None and int_file_df is not None:
        ova_id_name = ova_id
        int_id_name = int_id
        ova_file_df[ova_id_name] = ova_file_df[ova_id_name].astype(str)
        int_file_df[int_id_name] = int_file_df[int_id_name].astype(str)

        missing_int_tx = get_missing_tx(
            x=ova_file_df[ova_id_name].astype("string"),
            y=int_file_df[int_id_name].astype("string"),
            alt_x=ova_file_df[alt_ova_id].astype("string"),
            alt_y=int_file_df[alt_int_id].astype("string"),
        ).values

        missing_ova_tx = get_missing_tx(
            x=int_file_df[int_id_name].astype("string"),
            y=ova_file_df[ova_id_name].astype("string"),
            alt_x=int_file_df[alt_int_id].astype("string"),
            alt_y=ova_file_df[alt_ova_id].astype("string"),
        ).values

        int_amount_col = ""
        for name in amount_col_names:
            if name in int_file_df.columns:
                if not int_file_df[name].isna().all():
                    int_amount_col = name
                    break

        ova_amount_col = ""
        for name in amount_col_names:
            if name in ova_file_df.columns:
                if not ova_file_df[name].isna().all():
                    ova_amount_col = name
                    break
        ova_file_df[ova_amount_col] = ova_file_df[ova_amount_col].astype("float")
        int_file_df[int_amount_col] = int_file_df[int_amount_col].astype("float")
        missing_ova_amount_name = ova_amount_col
        missing_int_amount_name = int_amount_col

        missing_ova_data = int_file_df[
            int_file_df[int_id_name].astype("string").isin(missing_ova_tx)
            | int_file_df[alt_int_id].astype("string").isin(missing_ova_tx)
        ]
        missing_ova_value = missing_ova_data[missing_int_amount_name].abs().sum()

        missing_int_data = ova_file_df[
            ova_file_df[ova_id_name].astype("string").isin(missing_int_tx)
            | ova_file_df[alt_ova_id].astype("string").isin(missing_int_tx)
        ]

        missing_int_value = missing_int_data[missing_ova_amount_name].abs().sum()

        write_missing_ova_data(
            amount_col_name=int_amount_col,
            df=missing_ova_data,
            file_name=recons_file,
            value=missing_ova_value,
        )
        write_missing_int_data(
            amount_col_name=ova_amount_col,
            df=missing_int_data,
            file_name=recons_file,
            value=missing_int_value,
        )


def gip_custom(
    ova_files: Tuple[str or None, str or None],
    int_file: str or None,
    num_lines_of_header: Tuple[int, int],
    alt_recons_name: str,
    file_output_name: str,
):
    # ------------------------ OVA ------------------------
    recons_file = (
        f"{alt_recons_name} - Recons.xlsx"
        if ova_files[0] is None
        else f"{ova_files[0][:-5]} - Recons.xlsx"
    )
    ova_files = tuple(item for item in ova_files if item is not None)
    if len(ova_files) == 2:
        ova_df1 = pd.read_excel(ova_files[0])
        ova_df2 = pd.read_excel(ova_files[1])

        ova_volume = len(ova_df1) + len(ova_df2)
        ova_volumes[10] = ova_volume
        amount_col = ""
        for name in amount_col_names:
            if name in ova_df1.columns:
                if not ova_df1[name].isna().all():
                    amount_col = name
                    break
        ova_value = ova_df1[amount_col].abs().sum() + ova_df2[amount_col].abs().sum()
        ova_values[10] = ova_value
        print(f"GIP OVA VOLUME : {ova_volume}")
        print(f"GIP OVA VALUE : {ova_value}")

        with pd.ExcelWriter(
            recons_file, engine="openpyxl", mode="w"
        ) as writer:  # specify new file name to write to
            ova_df1.to_excel(
                writer, sheet_name="Sheet1", index=False
            )  # save original data into first sheet of new file
    # ---------------------------- INT -------------------------
    if int_file is not None:
        int_df = pd.read_excel(int_file)
        int_volume = len(int_df)
        int_volumes[10] = int_volume
        amount_col = ""
        for name in amount_col_names:
            if name in int_df.columns:
                if not int_df[name].isna().all():
                    amount_col = name
                    break

        int_value = int_df[amount_col].abs().sum()
        int_values[10] = int_value
        print(f"GIP INT VOLUME {int_volume}")
        print(f"GIP INT VALUE {int_value}")
        dup, dup_val = find_duplicates(int_df)
        write_duplicate_data(
            amount_col_name=amount_col, df=dup, value=dup_val, file_name=recons_file
        )


def get_missing_tx(
    x: pd.Series,
    y: pd.Series,
    alt_x: pd.Series,
    alt_y: pd.Series,
) -> pd.Series:
    """
    Returns the elements in Series 'x' that are not present in Series 'y'.

    Parameters:
    ----------
    x (pd.Series): A pandas Series containing elements to be checked for presence in 'y'.
    y (pd.Series): A pandas Series containing elements to be checked against for presence.

    Returns:
    -------
    pd.Series: A pandas Series containing elements from 'x' that are missing in 'y'.

    Note:
    ----
    - This function performs a check to find elements in Series 'x' that are not present in Series 'y'.
    - The resulting Series contains only the elements that are missing in 'y' while maintaining the original order from 'x'.
    """
    x = remove_leading_zeros(x)
    y = remove_leading_zeros(y)

    missing = (~x.astype(str).str.lower().isin(y.astype(str).str.lower())) & (
        ~alt_x.astype(str).str.lower().isin(alt_y.astype(str).str.lower())
    )
    x_missing = x[missing].combine_first(alt_x[missing])
    x_missing[(x_missing == "nan")] = alt_x[missing]
    return x_missing


def remove_leading_zeros(series):
    # Remove leading zeros from the series
    return series.str.replace(r"^0+", "", regex=True)


if __name__ == "__main__":
    prompt = input("Recons for yesterday? (Y/N) :")
    if prompt.upper() == "Y":
        date_ = date.today() - timedelta(1)
    else:
        print("Please enter the date...")
        day = int(input("Recons Day (1-31): "))
        month = int(input("Recons Month (1-12): "))
        year = int(input("Recons Year (eg 2023): "))
        date_ = datetime(year=year, month=month, day=day)

    yesterday = f"_{get_date(date_, format='normal')}"
    print(yesterday)

    recons_yesterday = f"{get_date(date_, format='recons')} 00:00:00"
    print(recons_yesterday)

    current_month = date_.strftime("%b").upper()  # JAN, FEB
    print(current_month)

    GIPdate = get_date(date=date_ + timedelta(1), format="gip")
    print(GIPdate)
    ova_volumes = [0] * 17
    ova_values = [0] * 17
    int_volumes = [0] * 17
    int_values = [0] * 17
    dup_volumes = [0] * 17
    dup_values = [0.00] * 17
    list_index = 0
    run_recons(
        (
            check_for_file(f"Ngenius{yesterday}.xlsx"),
            check_for_file(f"Ngenius KC{yesterday}.xlsx"),
        ),
        num_lines_of_header=(0, 0),
        alt_recons_name=f"Ngenius{yesterday}",
        file_output_name="Ngenius",
        ova_status_flag="SUCCESS",
        ova_status_col="Payment Status",
        list_index=1,
        ova_id="Merchant Defined Order Number",
        int_id="Universal Transaction Reference",
        alt_int_id="ID",
        alt_ova_id="System Generated Order",
    )
    try:
        run_recons(
            (
                check_for_file(f"KR MTN Credit{yesterday}.xlsx"),
                check_for_file(f"KR MTN Disb_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(0, 0),
            alt_recons_name=f"KR MTN Credit{yesterday}",
            file_output_name="MTN_KR_Credit",
            list_index=2,
            ova_id="External Transaction Id",
            int_id="Integrator Trans ID",
            alt_int_id="Bill Er Trans ID",
            alt_ova_id="Id",
        )
    except:
        ova_volumes[list_index] = 0
        int_volumes[list_index] = 0
        int_values[list_index] =0
        int_volumes[list_index] = 0
        run_recons(
            (
                check_for_file(f"KR MTN Credit{yesterday}.xlsx"),
                check_for_file(f"KR MTN Disb_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(0, 0),
            alt_recons_name=f"KR MTN Credit{yesterday}",
            file_output_name="MTN_KR_Credit",
            list_index=2,
            ova_id="External id",
            int_id="IntegratorTransId",
            alt_int_id="BillerTransId",
            alt_ova_id="Id",
        )
    try:
        run_recons(
            (
                check_for_file(f"KR MTN Debit{yesterday}.xlsx"),
                check_for_file(f"KR MTN Coll_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(0, 0),
            alt_recons_name=f"KR MTN Debit{yesterday}",
            file_output_name="MTN_KR_Debit",
            list_index=3,
            ova_id="External Transaction Id",
            int_id="Integrator Trans ID",
            alt_ova_id="Id",
            alt_int_id="Bill Er Trans ID",
        )
    except:
        ova_volumes[list_index] = 0
        int_volumes[list_index] = 0
        int_values[list_index] =0
        int_volumes[list_index] = 0
        run_recons(
            (
                check_for_file(f"KR MTN Debit{yesterday}.xlsx"),
                check_for_file(f"KR MTN Coll_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(0, 0),
            alt_recons_name=f"KR MTN Debit{yesterday}",
            file_output_name="MTN_KR_Debit",
            list_index=3,
            ova_id="External id",
            int_id="IntegratorTransId",
            alt_ova_id="Id",
            alt_int_id="BillerTransId",
        )
    run_recons(
         (
             check_for_file(f"KR AT{yesterday}.xlsx"),
             check_for_file(f"KR AT Coll_mBase{yesterday}.xlsx"),
         ),
         num_lines_of_header=(0, 0),
         alt_recons_name=f"AT Cashin{yesterday}",
         file_output_name="AT_KR_Cashin",
         ova_status_flag="Merchant Payment",
         ova_status_col="Service Type",
         mb_status_flag="CONFIRMED",
         list_index=4,
         ova_id="External Transaction Id",
         int_id="Transaction Id",
         alt_int_id="",
         alt_ova_id="",
     )
    run_recons(
        (
            check_for_file(f"KR AT{yesterday}.xlsx"),
            check_for_file(f"KR AT Disb_mBase{yesterday}.xlsx"),
        ),
        num_lines_of_header=(0, 0),
        alt_recons_name=f"AT Cashout{yesterday}",
        file_output_name="AT_KR_Cashout",
        ova_status_flag="Cash in",
        ova_status_col="Service Type",
        mb_status_flag="CONFIRMED",
        list_index=5,
        ova_id="External Transaction Id",
        int_id="Transaction Id",
        alt_int_id="Transaction Id",
        alt_ova_id="External Transaction Id",
    )
    try:
        run_recons(
            (
                check_for_file(f"KR Telecel Cashin{yesterday}.xlsx"),
                check_for_file(f"KR Telecel Coll_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(5, 0),
            alt_recons_name=f"KR Telecel Cashin{yesterday}",
            file_output_name="Telecel KR Cashin",
            mb_status_flag="CONFIRMED",
            list_index=6,
            ova_id="TransId",
            int_id="Integrator Trans ID",
            alt_int_id="Bill Er Trans ID",
            alt_ova_id="Receipt No.",
        )
    except:
        ova_volumes[list_index] = 0
        int_volumes[list_index] = 0
        int_values[list_index] =0
        int_volumes[list_index] = 0
        run_recons(
            (
                check_for_file(f"KR Telecel Cashin{yesterday}.xlsx"),
                check_for_file(f"KR Telecel Coll_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(5, 0),
            alt_recons_name=f"KR Telecel Cashin{yesterday}",
            file_output_name="Telecel KR Cashin",
            mb_status_flag="CONFIRMED",
            list_index=6,
            ova_id="TransId",
            int_id="Transaction Id",
            alt_int_id="Receipt No",
            alt_ova_id="Receipt No.",
        )

    try:
        run_recons(
            (
                check_for_file(f"KR Telecel Cashout{yesterday}.xlsx"),
                check_for_file(f"KR Telecel Disb_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(5, 0),
            alt_recons_name=f"KR Telecel Cashout{yesterday}",
            file_output_name="Telecel KR Cashout",
            mb_status_flag="CONFIRMED",
            list_index=7,
            ova_id="TransId",
            int_id="Integrator Trans ID",
            alt_int_id="Bill Er Trans ID",
            alt_ova_id="Receipt No.",
        )
    except:
        ova_volumes[list_index] = 0
        int_volumes[list_index] = 0
        int_values[list_index] =0
        int_volumes[list_index] = 0
        run_recons(
            (
                check_for_file(f"KR Telecel Cashout{yesterday}.xlsx"),
                check_for_file(f"KR Telecel Disb_mBase{yesterday}.xlsx"),
            ),
            num_lines_of_header=(5, 0),
            alt_recons_name=f"KR Telecel Cashout{yesterday}",
            file_output_name="Telecel KR Cashout",
            mb_status_flag="CONFIRMED",
            list_index=7,
            ova_id="TransId",
            int_id="Transaction Id",
            alt_int_id="Receipt No",
            alt_ova_id="Receipt No.",
        )


    run_recons(
        (
            check_for_file(f"Npontu MTN Credit{yesterday}.xlsx"),
            check_for_file(f"Npontu MTN Coll_mBase{yesterday}.xlsx"),
        ),
        num_lines_of_header=(0, 0),
        alt_recons_name=f"Npontu MTN Credit{yesterday}",
        file_output_name="Npontu MTN Credit",
        mb_status_flag="CONFIRMED",
        list_index=8,
        ova_id="External id",
        int_id="IntegratorTransId",
        alt_int_id="BillerTransId",
        alt_ova_id="Id",
    )

    run_recons(
        (
            check_for_file(f"Npontu MTN Debit{yesterday}.xlsx"),
            check_for_file(f"Npontu MTN Disb_mBase{yesterday}.xlsx"),
        ),
        num_lines_of_header=(0, 0),
        alt_recons_name=f"Npontu MTN Debit{yesterday}",
        file_output_name="Npontu MTN Debit",
        mb_status_flag="CONFIRMED",
        list_index=9,
        ova_id="External id",
        int_id="IntegratorTransId",
        alt_int_id="BillerTransId",
        alt_ova_id="Id",
    )
    run_recons(
        (
            check_for_file(f"GIP{yesterday}.xlsx"),
            check_for_file(f"GIP Metabase{yesterday}.xlsx"),
        ),
        num_lines_of_header=(0, 0),
        file_output_name="GIP",
        ova_id="REFERENCE_NUMBER",
        int_id="IntegratorTransId",
        list_index=10,
        alt_recons_name=f"GIP_{yesterday}",
        alt_ova_id="REFERENCE_NUMBER",
        alt_int_id="IntegratorTransId"

    )
    # gip_custom(
    #     ova_files=(
    #         check_for_file(f"slydepay_sending_{GIPdate}'.xlsx"),
    #         check_for_file(f"slydepay_sendingGhlink_{GIPdate}.xlsx"),
    #     ),
    #     num_lines_of_header=(0, 0),
    #     alt_recons_name=f"slydepay_sending_{yesterday}",
    #     file_output_name="GIP",
    #     int_file=check_for_file(f"GIP Metabase{yesterday}.xlsx"),
    # )


update_recons_sheet()
    # run_recons(
    #     (
    #         check_for_file(f"Quipu{yesterday}.xlsx"),
    #         check_for_file(f"Quipu KC{yesterday}.xlsx"),
    #     ),
    #     num_lines_of_header=(0, 0),
    #     alt_recons_name=f"Quipu{yesterday}",
    #     file_output_name="QUIPU",
    #     ova_status_flag="SUCCESS",
    #     ova_status_col="Status",
    #     list_index=23,
    #     ova_id="Order Code",
    #     int_id="Universal Transaction Reference",
    #     alt_int_id="Receipt No",
    #     alt_ova_id="Order Code",
# run_recons(
#     (
#         check_for_file(f"MPGS{yesterday}.xlsx"),
#         check_for_file(f"MPGS_trn{yesterday}.xlsx"),
#     ),
#     num_lines_of_header=(0, 0),
#     alt_recons_name=f"MPGS{yesterday}",
#     file_output_name="MPGS",
#     mb_status_flag="CONFIRMED",
#     alt_int_id="Transaction Id",
#     alt_ova_id="Order ID",
#     list_index=22,
#     ova_id="Order ID",
#     int_id="Transaction Id",
# )
    # )
    # run_recons(
    #     (
    #         check_for_file(f"MIGS 01{yesterday}.xlsx"),
    #         check_for_file(f"MIGS 01 Metabase{yesterday}.xlsx"),
    #     ),
    #     num_lines_of_header=(3, 0),
    #     alt_recons_name=f"MIGS 01{yesterday}",
    #     file_output_name="MIGS_01",
    #     list_index=0,
    #     ova_id="Merchant Transaction Reference",
    #     int_id="External Payment Request â†’ Institution Trans ID",
    #     alt_int_id="Institution Trans ID",
    #     alt_ova_id="Transaction ID",
    # )

    # try:
    #     run_recons(
    #         (
    #             check_for_file(f"MPGS{yesterday}.xlsx"),
    #             check_for_file(f"MPGS_trn{yesterday}.xlsx"),
    #         ),
    #         num_lines_of_header=(0, 0),
    #         alt_recons_name=f"MPGS{yesterday}",
    #         file_output_name="MPGS",
    #         mb_status_flag="CONFIRMED",
    #         alt_int_id="BillerTransId",
    #         alt_ova_id="Acquirer Transaction ID",
    #         list_index=0,
    #         ova_id="Acquirer Transaction ID",
    #         int_id="BillerTransId",
    #     )
    # except:
    #     run_recons(
    #         (
    #             check_for_file(f"MPGS{yesterday}.xlsx"),
    #             check_for_file(f"MPGS_trn{yesterday}.xlsx"),
    #         ),
    #         num_lines_of_header=(0, 0),
    #         alt_recons_name=f"MPGS{yesterday}",
    #         file_output_name="MPGS",
    #         mb_status_flag="CONFIRMED",
    #         alt_int_id="Bill Er Trans ID",
    #         alt_ova_id="Acquirer Transaction ID",
    #         list_index=0,
    #         ova_id="Acquirer Transaction ID",
    #         int_id="Bill Er Trans ID",
    #     )
    
    # run_recons(
    #     (
    #         check_for_file(f"MTN Prompt{yesterday}.xlsx"),
    #         check_for_file(f"Metabase{yesterday}.xlsx"),
    #     ),
    #     num_lines_of_header=(0, 0),
    #     mb_service_name="MTN Money MADAPI",
    #     mb_creditDebit_flag="C",
    #     alt_recons_name=f"MTN Prompt{yesterday}",
    #     file_output_name="MTN Prompt",
    #     list_index=1,
    #     ova_id="External Transaction Id",
    #     int_id="IntegratorTransId",
    #     alt_int_id="BillerTransId",
    #     alt_ova_id="Id",
    # )

    # run_recons(
    #     (
    #         check_for_file(f"MTN Cashout{yesterday}.xlsx"),
    #         check_for_file(f"Metabase{yesterday}.xlsx"),
    #     ),
    #     num_lines_of_header=(0, 0),
    #     alt_recons_name=f"MTN Cashout{yesterday}",
    #     file_output_name="MTN_PORTAL",
    #     mb_service_name="MTN Money MADAPI",
    #     mb_creditDebit_flag="D",
    #     list_index=3,
    #     ova_id="External Transaction Id",
    #     int_id="IntegratorTransId",
    #     alt_int_id="BillerTransId",
    #     alt_ova_id="Id",
    # )
    # run_recons(
    #     (
    #         check_for_file(f"AT Cashout{yesterday}.xlsx"),
    #         check_for_file(f"Metabase{yesterday}.xlsx"),
    #     ),
    #     num_lines_of_header=(4, 0),
    #     alt_recons_name=f"AT Cashout{yesterday}",
    #     file_output_name="AT_CASHOUT",
    #     mb_service_name="ATMoney_Slydepay",
    #     mb_creditDebit_flag="D",
    #     list_index=5,
    #     ova_id="Transaction Id",
    #     int_id="IntegratorTransId",
    #     alt_int_id="integratorTransId",
    #     alt_ova_id="Transaction Id",
    # )
    # run_recons(
    #     (
    #         check_for_file(f"Telecel Cashin{yesterday}.xlsx"),
    #         check_for_file(f"Metabase{yesterday}.xlsx"),
    #     ),
    #     num_lines_of_header=(5, 0),
    #     alt_recons_name=f"Telecel Cashin{yesterday}",
    #     file_output_name="Telecel CASHIN",
    #     mb_service_name="Telecel Cash",
    #     mb_creditDebit_flag="C",
    #     list_index=6,
    #     ova_id="TransId",
    #     int_id="IntegratorTransId",
    #     alt_int_id="BillerTransId",
    #     alt_ova_id="Receipt No.",
    # )
    # run_recons(
    #     (
    #         check_for_file(f"Telecel Cashout{yesterday}.xlsx"),
    #         check_for_file(f"Metabase{yesterday}.xlsx"),
    #     ),
    #     num_lines_of_header=(5, 0),
    #     alt_recons_name=f"Telecel Cashout{yesterday}",
    #     file_output_name="Telecel CASHOUT",
    #     mb_creditDebit_flag="D",
    #     mb_service_name="Telecel Cash",
    #     list_index=7,
    #     ova_id="TransId",
    #     int_id="IntegratorTransId",
    #     alt_int_id="BillerTransId",
    #     alt_ova_id="Receipt No.",
    # )
    # run_recons(
    #     (
    #         check_for_file(f"Stanbic FI Credit{yesterday}.xlsx"),
    #         check_for_file(f"Stanbic FI Credit Metabase{yesterday}.xlsx"),
    #     ),
    #     num_lines_of_header=(0, 0),
    #     alt_recons_name=f"Stanbic FI Credit{yesterday}",
    #     file_output_name="Stanbic FI CREDIT",
    #     mb_status_flag="CONFIRMED",
    #     list_index=8,
    #     ova_id="REMARKS2",
    #     int_id="External Payment Request → Institution Trans ID",
    #     alt_int_id="External Payment Request → Institution Trans ID",
    #     alt_ova_id="REMARKS2",
    # )
    # run_recons(
    #     (
    #         check_for_file(f"MIGS08{yesterday}.xlsx"),
    #         check_for_file(f"MiGS_trn{yesterday}.xlsx"),
    #     ),
    #     num_lines_of_header=(3, 0),
    #     alt_recons_name=f"MIGS 08{yesterday}",
    #     file_output_name="BB MIG",
    #     ova_id="Order ID",
    #     int_id="Receipt No",
    #     alt_ova_id="Order ID",
    #     alt_int_id="Receipt No",
    #     mb_status_flag="CONFIRMED",
    #     list_index=21,
    # )